from __future__ import annotations

import math

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference, Series
from openpyxl.chart.data_source import NumDataSource, NumRef
from openpyxl.chart.error_bar import ErrorBars
from openpyxl.utils import get_column_letter

from translations import t

_LEFT_COLOUR  = "E74C3C"   # red, no leading #
_RIGHT_COLOUR = "2ECC71"   # green


def export_excel(processed: dict, out_path: str) -> None:
    """
    Export ROM analysis results to an xlsx file.

    Sheet 1 — Summary: wide-format ROM table + native Excel bar chart.
    Sheet 2 — Segments: per-repetition data split into Left and Right tables.

    Args:
        processed: {(movement_name, side): data_dict} from the GUI session.
        out_path:  Destination .xlsx path.
    """
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Summary"
    ws2 = wb.create_sheet("Segments")

    _write_sheet1(ws1, processed)
    _write_sheet2(ws2, processed)

    wb.save(out_path)


# ── Sheet 1 ───────────────────────────────────────────────────────────────────

def _write_sheet1(ws, processed: dict) -> None:
    # Collect unique movements in first-appearance order
    movements: list[str] = []
    seen: set[str] = set()
    for mv_name, _side in processed:
        if mv_name not in seen:
            movements.append(mv_name)
            seen.add(mv_name)

    sides_present = {side for _mv, side in processed}
    has_right = "Right" in sides_present
    has_left  = bool(sides_present - {"Right"})   # anything that's not "Right"

    # ── Title ─────────────────────────────────────────────────────────────────
    ws.cell(row=1, column=1, value=t("rom_summary_chart_title"))

    # ── Column headers (row 3) ────────────────────────────────────────────────
    # Wide format: one row per movement, Left stats in cols 2-6, Right in 7-11
    HDR_ROW = 3
    headers = [
        t("s4_hdr_movement"),
        "N (L)",
        "Mean L (°)",
        "SD L (°)",
        "Min L (°)",
        "Max L (°)",
        "N (R)",
        "Mean R (°)",
        "SD R (°)",
        "Min R (°)",
        "Max R (°)",
    ]
    for col, h in enumerate(headers, 1):
        ws.cell(row=HDR_ROW, column=col, value=h)

    # ── Data rows ─────────────────────────────────────────────────────────────
    DATA_START = HDR_ROW + 1
    for i, mv_name in enumerate(movements):
        row = DATA_START + i
        ws.cell(row=row, column=1, value=mv_name)

        # Left data: side "Left" or "—" (no-prefix unilateral)
        left_key = next(
            ((mv_name, s) for s in ("Left", "—") if (mv_name, s) in processed),
            None,
        )
        if left_key:
            _write_rom_stats(ws, row, 2, processed[left_key])

        # Right data
        if (mv_name, "Right") in processed:
            _write_rom_stats(ws, row, 7, processed[(mv_name, "Right")])

    last_data_row = DATA_START + len(movements) - 1

    # ── Native bar chart ──────────────────────────────────────────────────────
    chart = BarChart()
    chart.type = "col"
    chart.grouping = "clustered"
    chart.title = t("rom_summary_chart_title")
    chart.y_axis.title = "ROM (°)"
    chart.width  = 15   # cm
    chart.height = 10   # cm

    cats = Reference(
        ws,
        min_col=1, max_col=1,
        min_row=DATA_START, max_row=last_data_row,
    )

    if has_left:
        mean_ref = Reference(ws, min_col=3, max_col=3,
                             min_row=DATA_START, max_row=last_data_row)
        series_l = Series(mean_ref, title="Left")
        try:
            series_l.graphicalProperties.solidFill = _LEFT_COLOUR
        except Exception:
            pass
        try:
            series_l.errBars = _make_error_bars(
                ws.title, col=4, first_row=DATA_START, last_row=last_data_row)
        except Exception:
            pass
        chart.append(series_l)

    if has_right:
        mean_ref = Reference(ws, min_col=8, max_col=8,
                             min_row=DATA_START, max_row=last_data_row)
        series_r = Series(mean_ref, title="Right")
        try:
            series_r.graphicalProperties.solidFill = _RIGHT_COLOUR
        except Exception:
            pass
        try:
            series_r.errBars = _make_error_bars(
                ws.title, col=9, first_row=DATA_START, last_row=last_data_row)
        except Exception:
            pass
        chart.append(series_r)

    chart.set_categories(cats)
    ws.add_chart(chart, f"A{last_data_row + 3}")


def _write_rom_stats(ws, row: int, col_start: int, data: dict) -> None:
    """Write N, Mean, SD, Min, Max ROM stats starting at col_start."""
    stats = data.get("extended", {}).get("rom", {})
    values = stats.get("values", [])
    mean_v = stats.get("mean", float("nan"))
    sd_v   = stats.get("sd",   0.0)
    min_v  = stats.get("min",  float("nan"))
    max_v  = stats.get("max",  float("nan"))

    ws.cell(row=row, column=col_start,     value=len(values))
    ws.cell(row=row, column=col_start + 1, value=_safe_float(mean_v))
    ws.cell(row=row, column=col_start + 2, value=round(sd_v, 2))
    ws.cell(row=row, column=col_start + 3, value=_safe_float(min_v))
    ws.cell(row=row, column=col_start + 4, value=_safe_float(max_v))


# ── Sheet 2 ───────────────────────────────────────────────────────────────────

def _write_sheet2(ws, processed: dict) -> None:
    # Split processed entries by side
    left_entries  = [
        (mv, data) for (mv, side), data in processed.items()
        if side != "Right"
    ]
    right_entries = [
        (mv, data) for (mv, side), data in processed.items()
        if side == "Right"
    ]

    col_headers = [
        t("frame"),
        t("angle_deg"),
        "Rep",
        t("s4_hdr_movement"),
        "ROM (°)",
    ]

    row = 1
    row = _write_side_table(ws, t("left_side"),  left_entries,  col_headers, row)
    row += 2   # two blank rows between tables
    _write_side_table(ws, t("right_side"), right_entries, col_headers, row)


def _write_side_table(
    ws,
    side_label: str,
    entries: list[tuple[str, dict]],
    col_headers: list[str],
    start_row: int,
) -> int:
    """
    Write one side's table.  Returns the next available row after all data.
    Column order: Frame | Angle (°) | Rep | Movement | ROM (°)
    """
    row = start_row

    # Section header
    ws.cell(row=row, column=1, value=side_label)
    row += 1

    # Column headers
    for col, h in enumerate(col_headers, 1):
        ws.cell(row=row, column=col, value=h)
    row += 1

    if not entries:
        ws.cell(row=row, column=1, value=t("no_data"))
        return row + 1

    wrote_any = False
    for mv_name, data in entries:
        angle_arr = data.get("angle_data")
        segments  = data.get("segments", [])
        rom_vals  = data.get("extended", {}).get("rom", {}).get("values", [])

        if segments:
            # Continuous mode: one row per segment (repetition)
            for rep_idx, (s, e) in enumerate(segments):
                frame = int(s)
                if angle_arr is not None and frame < len(angle_arr):
                    angle_val = float(angle_arr[frame])
                    angle_cell = round(angle_val, 2) if not math.isnan(angle_val) else None
                else:
                    angle_cell = None

                rom_val = rom_vals[rep_idx] if rep_idx < len(rom_vals) else float("nan")

                ws.cell(row=row, column=1, value=frame)
                ws.cell(row=row, column=2, value=angle_cell)
                ws.cell(row=row, column=3, value=rep_idx + 1)
                ws.cell(row=row, column=4, value=mv_name)
                ws.cell(row=row, column=5, value=_safe_float(rom_val))
                row += 1
                wrote_any = True
        else:
            # Individual mode: per-rep stats without frame numbers
            for rep_idx, rom_val in enumerate(rom_vals):
                ws.cell(row=row, column=1, value=None)
                ws.cell(row=row, column=2, value=None)
                ws.cell(row=row, column=3, value=rep_idx + 1)
                ws.cell(row=row, column=4, value=mv_name)
                ws.cell(row=row, column=5, value=_safe_float(rom_val))
                row += 1
                wrote_any = True

    if not wrote_any:
        ws.cell(row=row, column=1, value=t("no_data"))
        row += 1

    return row


# ── Helpers ───────────────────────────────────────────────────────────────────

def _safe_float(v: float) -> float | None:
    """Return rounded float or None for NaN."""
    return round(v, 2) if not math.isnan(v) else None


def _make_error_bars(
    sheet_title: str,
    col: int,
    first_row: int,
    last_row: int,
) -> ErrorBars:
    """Build ±SD custom error bars referencing a worksheet column."""
    col_letter = get_column_letter(col)
    range_str = f"'{sheet_title}'!${col_letter}${first_row}:${col_letter}${last_row}"

    nr = NumRef()
    nr.f = range_str

    nd = NumDataSource()
    nd.numRef = nr

    eb = ErrorBars()
    eb.errDir    = "y"
    eb.errBarType = "both"
    eb.errValType = "cust"
    eb.plus  = nd
    eb.minus = nd
    return eb
