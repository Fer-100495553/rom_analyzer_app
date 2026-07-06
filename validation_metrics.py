from __future__ import annotations

import logging

import numpy as np
import pandas as pd

logger = logging.getLogger(__name__)


# ── Original C3D-based helpers (unchanged) ────────────────────────────────────

def extract_diff_vector(c3d_path: str, label: str) -> tuple[np.ndarray, float]:
    """
    Extracts a 3D difference vector from a C3D file.

    Points shape: (4, n_labels, n_frames) — rows 0=X, 1=Y, 2=Z, 3=residual.
    Subject prefix (e.g. 'Subject01:Diff_T8F_T8V') is stripped before matching.
    Frames where all three components are zero or residual < 0 are set to NaN.

    Returns:
        diff: np.ndarray shape (3, n_frames), dx dy dz in mm, invalid frames as NaN
        frame_rate: float
    """
    import ezc3d

    c3d = ezc3d.c3d(c3d_path)
    points = c3d["data"]["points"]  # (4, n_labels, n_frames)
    raw_labels = c3d["parameters"]["POINT"]["LABELS"]["value"]

    clean_labels = [
        lbl.split(":")[-1] if ":" in lbl else lbl for lbl in raw_labels
    ]

    if label not in clean_labels:
        raise ValueError(
            f"Label '{label}' not found in C3D file '{c3d_path}'. "
            f"Available labels (stripped): {clean_labels}"
        )

    idx = clean_labels.index(label)
    try:
        rate_val   = c3d["parameters"]["POINT"]["RATE"]["value"]
        frame_rate = float(rate_val[0] if hasattr(rate_val, "__len__") else rate_val)
    except (KeyError, IndexError, TypeError, ValueError):
        frame_rate = float(c3d["header"]["frame_rate"])

    x = points[0, idx, :].astype(float)
    y = points[1, idx, :].astype(float)
    z = points[2, idx, :].astype(float)
    residual = points[3, idx, :]

    diff = np.stack([x, y, z], axis=0)

    all_zero = (x == 0) & (y == 0) & (z == 0)
    invalid = all_zero | (residual < 0)
    diff[:, invalid] = np.nan

    return diff, frame_rate


def compute_euclidean_distance(diff: np.ndarray) -> np.ndarray:
    """
    Frame-by-frame Euclidean distance from a (3, n_frames) array.
    NaN frames propagate: any NaN component → NaN distance.
    Returns np.ndarray shape (n_frames,) in mm.
    """
    return np.sqrt(np.sum(diff ** 2, axis=0))


def compute_validation_metrics(distance: np.ndarray) -> dict:
    """
    Returns dict with keys:
        mean_mm, sd_mm, rmse_mm, max_mm, p95_mm, n_frames, n_valid
    """
    n_frames = len(distance)
    n_valid = int(np.sum(~np.isnan(distance)))
    return {
        "mean_mm": float(np.nanmean(distance)),
        "sd_mm":   float(np.nanstd(distance)),
        "rmse_mm": float(np.sqrt(np.nanmean(distance ** 2))),
        "max_mm":  float(np.nanmax(distance)),
        "p95_mm":  float(np.nanpercentile(distance, 95)),
        "n_frames": n_frames,
        "n_valid":  n_valid,
    }


def compute_per_axis_mae(diff: np.ndarray) -> dict:
    """Returns dict with keys: mae_x_mm, mae_y_mm, mae_z_mm."""
    return {
        "mae_x_mm": float(np.nanmean(np.abs(diff[0]))),
        "mae_y_mm": float(np.nanmean(np.abs(diff[1]))),
        "mae_z_mm": float(np.nanmean(np.abs(diff[2]))),
    }


# ── ROM statistical validation helpers ───────────────────────────────────────

def interpret_icc(icc_value: float) -> str:
    """
    Koo & Li (2016) ICC interpretation thresholds.

    Returns one of: 'poor', 'moderate', 'good', 'excellent'.
    """
    if icc_value < 0.50:
        return 'poor'
    if icc_value < 0.75:
        return 'moderate'
    if icc_value < 0.90:
        return 'good'
    return 'excellent'


def compute_metrics(y_ref: np.ndarray, y_alt: np.ndarray,
                    label_ref: str = 'ref', label_alt: str = 'alt') -> dict | None:
    """
    Compute agreement metrics between two paired ROM scalar arrays.

    Metrics:
        - Shapiro-Wilk W and p on the difference array (descriptive only).
        - Bland-Altman bias, SD of differences, 95% LoA (bias ± 1.96·SD).
        - RMSD.
        - ICC(3,1) with 95% CI via pingouin (two-way mixed, absolute agreement, single measures).

    Returns:
        dict with keys: n_reps, shapiro_W, shapiro_p, bias_deg, sd_diff_deg,
            loa_lower_deg, loa_upper_deg, rmsd_deg, icc, icc_ci95_lower,
            icc_ci95_upper, icc_interpretation.
        None if n < 3 (logs a warning).
    """
    try:
        import pingouin as pg
        from scipy.stats import shapiro
    except ImportError as exc:
        logger.error("compute_metrics requires 'pingouin' and 'scipy'. Install them first. %s", exc)
        return None

    n = len(y_ref)
    if n < 3:
        logger.warning(
            "compute_metrics: n=%d < 3 for %s vs %s. Skipping.", n, label_ref, label_alt
        )
        return None

    diff = np.asarray(y_alt, dtype=float) - np.asarray(y_ref, dtype=float)

    sw_stat, sw_p = shapiro(diff)

    bias      = float(np.mean(diff))
    sd_diff   = float(np.std(diff, ddof=1))
    loa_upper = bias + 1.96 * sd_diff
    loa_lower = bias - 1.96 * sd_diff
    rmsd      = float(np.sqrt(np.mean(diff ** 2)))

    df_long = pd.DataFrame({
        'targets': list(range(n)) * 2,
        'raters':  [label_ref] * n + [label_alt] * n,
        'ratings': list(y_ref)  + list(y_alt),
    })
    icc_result = pg.intraclass_corr(
        data=df_long, targets='targets', raters='raters', ratings='ratings'
    )
    row      = icc_result[icc_result['Type'] == 'ICC(C,1)'].iloc[0]
    icc_val  = float(row['ICC'])
    ci_lower = float(row['CI95'][0])
    ci_upper = float(row['CI95'][1])

    return {
        'n_reps':             n,
        'shapiro_W':          float(sw_stat),
        'shapiro_p':          float(sw_p),
        'bias_deg':           bias,
        'sd_diff_deg':        sd_diff,
        'loa_lower_deg':      loa_lower,
        'loa_upper_deg':      loa_upper,
        'rmsd_deg':           rmsd,
        'icc':                icc_val,
        'icc_ci95_lower':     ci_lower,
        'icc_ci95_upper':     ci_upper,
        'icc_interpretation': interpret_icc(icc_val),
    }


def run_comparison_A(rom_data: pd.DataFrame,
                     movement_list: list[str] | None = None,
                     movements_bilateral: list[str] | None = None) -> list[dict]:
    """
    Comparison A: effect of anatomical distance condition within each reconstructed model.

    Reference: (model, 'Perfect').
    Alternative: (model, anat_dist) for anat_dist in ['IJ_C7', 'PX_C7', 'IJ_PX_C7'].

    Args:
        rom_data: flat tidy DataFrame with columns Movement, Model, Anatomic_Distances,
                  Rep, Left_ROM_deg, Right_ROM_deg.
        movement_list: ordered list of all movements. If None, inferred from data.
        movements_bilateral: movements with no Right side. If None, inferred from NaN pattern.

    Returns:
        List of result dicts, each containing comparison='A', model, anat_dist,
        movement, side, and all keys from compute_metrics().
    """
    if movement_list is None:
        movement_list = sorted(rom_data['Movement'].unique())
    if movements_bilateral is None:
        movements_bilateral = (
            rom_data[rom_data['Right_ROM_deg'].isna()]['Movement'].unique().tolist()
        )

    def _sides(mv):
        return ['Left'] if mv in movements_bilateral else ['Left', 'Right']

    results = []
    for model in ['C7', 'NOT_C7']:
        for movement in movement_list:
            for side in _sides(movement):
                col  = f'{side}_ROM_deg'
                mask_ref = (
                    (rom_data['Movement'] == movement) &
                    (rom_data['Model']    == model) &
                    (rom_data['Anatomic_Distances'] == 'Perfect')
                )
                y_ref = rom_data[mask_ref].sort_values('Rep')[col].to_numpy(dtype=float)

                for anat_dist in ['IJ_C7', 'PX_C7', 'IJ_PX_C7']:
                    mask_alt = (
                        (rom_data['Movement'] == movement) &
                        (rom_data['Model']    == model) &
                        (rom_data['Anatomic_Distances'] == anat_dist)
                    )
                    y_alt = rom_data[mask_alt].sort_values('Rep')[col].to_numpy(dtype=float)
                    n = min(len(y_ref), len(y_alt))
                    m = compute_metrics(
                        y_ref[:n], y_alt[:n],
                        label_ref=f'{model}_Perfect',
                        label_alt=f'{model}_{anat_dist}',
                    )
                    if m is None:
                        continue
                    results.append({
                        'comparison': 'A',
                        'model':      model,
                        'anat_dist':  anat_dist,
                        'movement':   movement,
                        'side':       side,
                        **m,
                    })
    logger.info("run_comparison_A: produced %d result rows.", len(results))
    return results


def run_comparison_B(rom_data: pd.DataFrame,
                     movement_list: list[str] | None = None,
                     movements_bilateral: list[str] | None = None) -> list[dict]:
    """
    Comparison B: effect of kinematic model vs. Complete gold standard.

    Reference: ('Complete', 'Perfect').
    Alternative: (model, 'Perfect') for model in ['C7', 'NOT_C7'].

    Args:
        rom_data: flat tidy DataFrame (same format as run_comparison_A).
        movement_list: ordered list of all movements. If None, inferred from data.
        movements_bilateral: movements with no Right side. If None, inferred from NaN pattern.

    Returns:
        List of result dicts, each containing comparison='B', model, movement, side,
        and all keys from compute_metrics().
    """
    if movement_list is None:
        movement_list = sorted(rom_data['Movement'].unique())
    if movements_bilateral is None:
        movements_bilateral = (
            rom_data[rom_data['Right_ROM_deg'].isna()]['Movement'].unique().tolist()
        )

    def _sides(mv):
        return ['Left'] if mv in movements_bilateral else ['Left', 'Right']

    results = []
    for model in ['C7', 'NOT_C7']:
        for movement in movement_list:
            for side in _sides(movement):
                col = f'{side}_ROM_deg'
                mask_ref = (
                    (rom_data['Movement'] == movement) &
                    (rom_data['Model']    == 'Complete') &
                    (rom_data['Anatomic_Distances'] == 'Perfect')
                )
                mask_alt = (
                    (rom_data['Movement'] == movement) &
                    (rom_data['Model']    == model) &
                    (rom_data['Anatomic_Distances'] == 'Perfect')
                )
                y_ref = rom_data[mask_ref].sort_values('Rep')[col].to_numpy(dtype=float)
                y_alt = rom_data[mask_alt].sort_values('Rep')[col].to_numpy(dtype=float)
                n = min(len(y_ref), len(y_alt))
                m = compute_metrics(
                    y_ref[:n], y_alt[:n],
                    label_ref='Complete',
                    label_alt=model,
                )
                if m is None:
                    continue
                results.append({
                    'comparison': 'B',
                    'model':      model,
                    'movement':   movement,
                    'side':       side,
                    **m,
                })
    logger.info("run_comparison_B: produced %d result rows.", len(results))
    return results


def export_results_to_excel(results: dict[str, pd.DataFrame], output_path: str) -> None:
    """
    Export validation results to a formatted Excel workbook.

    Args:
        results: dict mapping sheet names to DataFrames. Expected keys:
                 'Comparison_A', 'Comparison_B', 'Summary_A', 'Summary_B'.
        output_path: destination .xlsx path (str or Path).

    Formatting applied:
        - Header row: bold, light blue fill, centered.
        - Numeric columns: 2 decimal places.
        - icc_interpretation column: colour-coded by quality tier.
        - Freeze top row on every sheet.
        - Auto-fit column widths.
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError as exc:
        logger.error("export_results_to_excel requires 'openpyxl'. %s", exc)
        return

    HEADER_FILL  = PatternFill('solid', fgColor='C9D6E8')
    HEADER_FONT  = Font(bold=True)
    HEADER_ALIGN = Alignment(horizontal='center')
    INTERP_FILLS = {
        'poor':      PatternFill('solid', fgColor='FFCCCC'),
        'moderate':  PatternFill('solid', fgColor='FFE0B2'),
        'good':      PatternFill('solid', fgColor='FFF9C4'),
        'excellent': PatternFill('solid', fgColor='C8E6C9'),
    }
    NUMERIC_COLS = {
        'n_reps', 'shapiro_W', 'shapiro_p',
        'bias_deg', 'sd_diff_deg', 'loa_lower_deg', 'loa_upper_deg',
        'rmsd_deg', 'icc', 'icc_ci95_lower', 'icc_ci95_upper',
    }

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    for sheet_name, df in results.items():
        ws = wb.create_sheet(sheet_name)
        headers = list(df.columns)
        ws.append(headers)
        for cell in ws[1]:
            cell.font      = HEADER_FONT
            cell.fill      = HEADER_FILL
            cell.alignment = HEADER_ALIGN
        ws.freeze_panes = 'A2'

        for row_vals in df.itertuples(index=False):
            ws.append(list(row_vals))

        for col_idx, col_name in enumerate(headers, start=1):
            col_letter = get_column_letter(col_idx)
            max_len = len(col_name)
            for row_idx in range(2, ws.max_row + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                val  = cell.value
                if col_name in NUMERIC_COLS and isinstance(val, (int, float)):
                    cell.number_format = '0.00'
                if col_name == 'icc_interpretation' and isinstance(val, str):
                    cell.fill = INTERP_FILLS.get(val, PatternFill())
                max_len = max(max_len, len(str(val)) if val is not None else 0)
            ws.column_dimensions[col_letter].width = max_len + 2

    wb.save(output_path)
    logger.info("export_results_to_excel: saved to %s", output_path)
