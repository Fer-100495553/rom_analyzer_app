from __future__ import annotations

import logging
import math
import os

import customtkinter as ctk
import numpy as np
import pandas as pd
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
from tkinter import filedialog, messagebox

import settings_manager
from translations import t

logger = logging.getLogger(__name__)

# ── Side display label for no-prefix variables ────────────────────────────
_NO_SIDE = "—"

# ── Widget keys excluded when saving/restoring import-row state ────────────
_ROW_WIDGET_KEYS = frozenset({
    "file_lbl", "status_lbl",
})


# ══════════════════════════════════════════════════════════════════════════
#  Small UI helpers
# ══════════════════════════════════════════════════════════════════════════

def _card(parent, title: str) -> ctk.CTkFrame:
    """Labelled card frame used as a section container."""
    frame = ctk.CTkFrame(parent)
    ctk.CTkLabel(
        frame, text=title,
        font=ctk.CTkFont(size=11, weight="bold"), text_color="gray",
    ).pack(anchor="w", padx=12, pady=(10, 4))
    return frame


def _display_name(mv_name: str) -> str:
    """Return the translated UI label for a movement, falling back to the key."""
    from config import MOVEMENT_DEFINITIONS
    key = MOVEMENT_DEFINITIONS.get(mv_name, {}).get("display_name_key", "")
    if not key:
        return mv_name
    result = t(key)
    return result if result != f"[{key}]" else mv_name


# ══════════════════════════════════════════════════════════════════════════
#  App — 4-screen ROM analysis application
# ══════════════════════════════════════════════════════════════════════════

class App(ctk.CTk):
    """
    Four-screen ROM analysis application.

    Screen 1 — Study Configuration: movement selection + laterality.
    Screen 2 — File Import: one C3D row per movement+side combination.
    Screen 3 — Segmentation: C3DSegmentationWindow per movement+side (modal).
    Screen 4 — Summary & Export: table, chart, CSV, report.
    """

    def __init__(self) -> None:
        super().__init__()
        self.title(t("app_title"))
        self.geometry("800x740")
        self.resizable(True, True)
        self.state("zoomed")
        self.bind("<F11>", lambda e: self.attributes("-fullscreen",
                                                     not self.attributes("-fullscreen")))
        self.bind("<Escape>", lambda e: self.attributes("-fullscreen", False))

        # ── Persistent session state ──────────────────────────────────────
        self._movement_vars: dict[str, ctk.BooleanVar] = {}
        self._laterality_var = ctk.StringVar(value="unilateral")
        self._recording_type_var = ctk.StringVar(value="continuous")
        self._num_reps_var = ctk.IntVar(value=6)
        self._import_rows: list[dict] = []
        self._process_btn: ctk.CTkButton | None = None
        self._processed: dict[tuple[str, str], dict] = {}
        self._layout_vertical: bool = False
        self._charts: dict = {}

        # Navigation state (used to re-render on language change)
        self._current_screen: int = 1
        self._selected_movements: list[str] = []

        # ── Persistent header ─────────────────────────────────────────────
        _hdr = ctk.CTkFrame(self, fg_color="transparent")
        _hdr.pack(fill="x", padx=20, pady=(20, 2))
        _hdr.grid_columnconfigure(0, weight=1)   # left spacer
        _hdr.grid_columnconfigure(1, weight=0)   # title (fixed width)
        _hdr.grid_columnconfigure(2, weight=1)   # right area (settings btn)

        self._title_lbl = ctk.CTkLabel(
            _hdr, text=t("app_title"),
            font=ctk.CTkFont(size=26, weight="bold"),
        )
        self._title_lbl.grid(row=0, column=0, columnspan=3, pady=(0, 2))

        self._subtitle_lbl = ctk.CTkLabel(
            _hdr, text=t("app_subtitle"), text_color="gray",
        )
        self._subtitle_lbl.grid(row=1, column=0, columnspan=3)

        # Settings button — top-right corner of the header
        _right_frame = ctk.CTkFrame(_hdr, fg_color="transparent")
        _right_frame.grid(row=0, column=2, rowspan=2, sticky="ne", pady=4)
        ctk.CTkButton(
            _right_frame, text="⚙", width=36, height=36,
            font=ctk.CTkFont(size=18),
            fg_color="transparent", border_width=1,
            command=self._open_settings,
        ).pack(anchor="e", padx=(0, 4))

        # Spacing after header
        ctk.CTkFrame(self, height=10, fg_color="transparent").pack()

        self._container = ctk.CTkFrame(self, fg_color="transparent")
        self._container.pack(fill="both", expand=True, padx=20, pady=(0, 18))

        self._show_screen_1()

    # ── Header refresh ─────────────────────────────────────────────────────

    def _refresh_header(self) -> None:
        self._title_lbl.configure(text=t("app_title"))
        self._subtitle_lbl.configure(text=t("app_subtitle"))
        self.title(t("app_title"))

    # ── Screen helpers ─────────────────────────────────────────────────────

    def _clear_container(self) -> None:
        for child in self._container.winfo_children():
            child.destroy()

    def _rerender_current_screen(self) -> None:
        if self._current_screen == 1:
            self._show_screen_1()
        elif self._current_screen == 2:
            self._show_screen_2(self._selected_movements)
        elif self._current_screen == 4:
            self._show_screen_4()

    # ── Settings dialog ────────────────────────────────────────────────────

    def _open_settings(self) -> None:
        dlg = ctk.CTkToplevel(self)
        dlg.title(t("settings_title"))
        dlg.geometry("460x290")
        dlg.minsize(460, 290)
        dlg.resizable(False, False)
        dlg.grab_set()
        dlg.lift()
        dlg.focus_force()

        # ── Language ──────────────────────────────────────────────────────
        ctk.CTkLabel(
            dlg, text=t("settings_language_label"),
            font=ctk.CTkFont(size=13, weight="bold"),
        ).pack(anchor="w", padx=28, pady=(24, 8))

        lang_frame = ctk.CTkFrame(dlg, fg_color="transparent")
        lang_frame.pack(anchor="w", padx=28, pady=(0, 4))

        lang_var = ctk.StringVar(value=settings_manager.get("language"))
        for label, value in [("English", "en"), ("Español", "es")]:
            ctk.CTkRadioButton(
                lang_frame, text=label,
                variable=lang_var, value=value,
                command=lambda v=value: self._on_language_change(v, dlg),
            ).pack(side="left", padx=(0, 28))

        # ── Theme ─────────────────────────────────────────────────────────
        ctk.CTkLabel(
            dlg, text=t("settings_theme_label"),
            font=ctk.CTkFont(size=13, weight="bold"),
        ).pack(anchor="w", padx=28, pady=(20, 8))

        theme_frame = ctk.CTkFrame(dlg, fg_color="transparent")
        theme_frame.pack(anchor="w", padx=28, pady=(0, 4))

        theme_var = ctk.StringVar(value=settings_manager.get("theme"))
        for label_key, value in [
            ("settings_theme_system", "System"),
            ("settings_theme_dark",   "Dark"),
            ("settings_theme_light",  "Light"),
        ]:
            ctk.CTkRadioButton(
                theme_frame, text=t(label_key),
                variable=theme_var, value=value,
                command=lambda v=value: self._on_theme_change(v),
            ).pack(side="left", padx=(0, 24))

        # ── Close ─────────────────────────────────────────────────────────
        ctk.CTkButton(
            dlg, text=t("settings_close"),
            width=120, command=dlg.destroy,
        ).pack(pady=(28, 18))

    def _on_language_change(self, lang: str, dlg: ctk.CTkToplevel) -> None:
        settings_manager.set_language(lang)
        dlg.title(t("settings_title"))
        self._refresh_header()
        self._rerender_current_screen()

    def _on_theme_change(self, theme: str) -> None:
        settings_manager.set_theme(theme)
        ctk.set_appearance_mode(theme)
        self._rerender_current_screen()

    # ══════════════════════════════════════════════════════════════════════
    #  Screen 1 — Study Configuration
    # ══════════════════════════════════════════════════════════════════════

    def _show_screen_1(self) -> None:
        self._current_screen = 1
        self._clear_container()
        f = ctk.CTkFrame(self._container, fg_color="transparent")
        f.pack(fill="both", expand=True)

        # ── A: Movement selection ─────────────────────────────────────────
        card_mv = _card(f, t("s1_card_movements"))
        card_mv.pack(fill="x", pady=(0, 8))

        from config import MOVEMENT_DEFINITIONS, MOVEMENT_PAIR_GROUPS
        self._movement_vars.clear()

        # Collect all movement names that belong to a pair group
        paired_members: set[str] = {
            mv for members in MOVEMENT_PAIR_GROUPS.values() for mv in members
        }
        # Map first member → (group_label, all_members) for insertion order
        pair_by_first: dict[str, tuple[str, list[str]]] = {}
        for group_label, members in MOVEMENT_PAIR_GROUPS.items():
            pair_by_first[members[0]] = (group_label, members)

        rendered_groups: set[str] = set()

        for mv_name, mv_def in MOVEMENT_DEFINITIONS.items():
            # Skip second (and further) pair members before creating any widget
            if mv_name in paired_members and mv_name not in pair_by_first:
                continue

            row = ctk.CTkFrame(card_mv, fg_color="transparent")
            row.pack(fill="x", padx=12, pady=2)

            if mv_name in paired_members:  # must be pair_by_first at this point
                group_label, members = pair_by_first[mv_name]
                if group_label not in rendered_groups:
                    rendered_groups.add(group_label)
                    group_var = ctk.BooleanVar(value=False)
                    for member in members:
                        self._movement_vars[member] = group_var
                    ctk.CTkCheckBox(
                        row, text=t("s1_mv_thorax_trunk_group"), variable=group_var,
                        width=380,
                    ).pack(side="left")
                continue

            var = ctk.BooleanVar(value=False)
            self._movement_vars[mv_name] = var

            is_optional = mv_def.get("optional", False)
            _dn = _display_name(mv_name)
            display = (_dn if not is_optional
                       else f"{_dn}  {t('s1_not_available')}")
            ctk.CTkCheckBox(
                row, text=display, variable=var, width=380,
                state="disabled" if is_optional else "normal",
            ).pack(side="left")

        sel_row = ctk.CTkFrame(card_mv, fg_color="transparent")
        sel_row.pack(anchor="w", padx=12, pady=(6, 10))
        ctk.CTkButton(sel_row, text=t("s1_select_all"), width=120,
                      command=self._select_all).pack(side="left", padx=(0, 8))
        ctk.CTkButton(sel_row, text=t("s1_deselect_all"), width=120,
                      command=self._deselect_all).pack(side="left")

        # ── B: Laterality ─────────────────────────────────────────────────
        card_lat = _card(f, t("s1_card_laterality"))
        card_lat.pack(fill="x", pady=(0, 8))

        lat_row = ctk.CTkFrame(card_lat, fg_color="transparent")
        lat_row.pack(anchor="w", padx=12, pady=(0, 10))

        for text_key, val, hint_key in [
            ("s1_unilateral", "unilateral", "s1_unilateral_hint"),
            ("s1_bilateral",  "bilateral",  "s1_bilateral_hint"),
        ]:
            col = ctk.CTkFrame(lat_row, fg_color="transparent")
            col.pack(side="left", padx=(0, 28))
            ctk.CTkRadioButton(
                col, text=t(text_key),
                variable=self._laterality_var, value=val,
            ).pack(anchor="w")
            ctk.CTkLabel(col, text=t(hint_key),
                         font=ctk.CTkFont(size=10), text_color="gray",
                         ).pack(anchor="w", padx=(22, 0))

        # ── C: Recording Type ─────────────────────────────────────────────
        card_rec = _card(f, t("s1_card_recording_type"))
        card_rec.pack(fill="x", pady=(0, 8))

        rec_row = ctk.CTkFrame(card_rec, fg_color="transparent")
        rec_row.pack(anchor="w", padx=12, pady=(0, 6))

        for text_key, val, hint_key in [
            ("s1_rec_continuous", "continuous", "s1_rec_continuous_hint"),
            ("s1_rec_individual", "individual", "s1_rec_individual_hint"),
        ]:
            col = ctk.CTkFrame(rec_row, fg_color="transparent")
            col.pack(side="left", padx=(0, 28))
            ctk.CTkRadioButton(
                col, text=t(text_key),
                variable=self._recording_type_var, value=val,
                command=self._on_recording_type_change,
            ).pack(anchor="w")
            ctk.CTkLabel(col, text=t(hint_key),
                         font=ctk.CTkFont(size=10), text_color="gray",
                         ).pack(anchor="w", padx=(22, 0))

        self._num_reps_row = ctk.CTkFrame(card_rec, fg_color="transparent")
        self._num_reps_row.pack(anchor="w", padx=12, pady=(0, 8))

        ctk.CTkLabel(
            self._num_reps_row, text=t("s1_num_reps_label"),
            font=ctk.CTkFont(size=11),
        ).pack(side="left", padx=(0, 8))

        ctk.CTkLabel(self._num_reps_row, text="  ", width=4).pack(side="left")
        ctk.CTkButton(
            self._num_reps_row, text="−", width=28, height=28,
            command=lambda: self._num_reps_var.set(
                max(1, self._num_reps_var.get() - 1)),
        ).pack(side="left")
        self._num_reps_lbl = ctk.CTkLabel(
            self._num_reps_row, textvariable=self._num_reps_var,
            width=32, font=ctk.CTkFont(size=13, weight="bold"),
        )
        self._num_reps_lbl.pack(side="left", padx=4)
        ctk.CTkButton(
            self._num_reps_row, text="+", width=28, height=28,
            command=lambda: self._num_reps_var.set(
                min(10, self._num_reps_var.get() + 1)),
        ).pack(side="left")

        self._on_recording_type_change()

        # ── D: Continue button ────────────────────────────────────────────
        ctk.CTkButton(
            f, text=t("s1_continue"),
            height=44, font=ctk.CTkFont(size=14, weight="bold"),
            command=self._go_to_screen_2,
        ).pack(fill="x", pady=(8, 0))

    def _select_all(self) -> None:
        from config import MOVEMENT_DEFINITIONS
        for mv_name, var in self._movement_vars.items():
            if not MOVEMENT_DEFINITIONS[mv_name].get("optional", False):
                var.set(True)

    def _deselect_all(self) -> None:
        for var in self._movement_vars.values():
            var.set(False)

    def _on_recording_type_change(self) -> None:
        is_individual = self._recording_type_var.get() == "individual"
        state = "normal" if is_individual else "disabled"
        for w in self._num_reps_row.winfo_children():
            try:
                w.configure(state=state)
            except Exception:
                pass

    def _go_to_screen_2(self) -> None:
        selected = [mv for mv, var in self._movement_vars.items() if var.get()]
        if not selected:
            messagebox.showerror(
                t("s1_err_no_movements_title"),
                t("s1_err_no_movements_msg"),
            )
            return
        # Clear state on real navigation (not language re-render)
        self._import_rows.clear()
        self._processed.clear()
        self._show_screen_2(selected)

    # ══════════════════════════════════════════════════════════════════════
    #  Screen 2 — File Import
    # ══════════════════════════════════════════════════════════════════════

    def _show_screen_2(self, selected_movements: list[str]) -> None:
        self._current_screen = 2
        self._selected_movements = selected_movements
        self._clear_container()

        is_individual = self._recording_type_var.get() == "individual"
        num_reps = self._num_reps_var.get() if is_individual else 1

        # Save loaded states before rebuilding widgets (language re-render)
        old_state: dict[tuple, dict] = {}
        for r in self._import_rows:
            key = (r["mv_name"], r["side"], r.get("rep_idx", 0))
            saved: dict = {k: v for k, v in r.items() if k not in _ROW_WIDGET_KEYS}
            old_state[key] = saved
        self._import_rows.clear()

        f = ctk.CTkFrame(self._container, fg_color="transparent")
        f.pack(fill="both", expand=True)

        card_title = (t("s2_card_import_individual") if is_individual
                      else t("s2_card_import"))
        card_imp = _card(f, card_title)
        card_imp.pack(fill="x", pady=(0, 8))

        # Table header
        hdr = ctk.CTkFrame(card_imp, fg_color="transparent")
        hdr.pack(fill="x", padx=12, pady=(0, 4))
        hdr_cols = [
            (t("s2_hdr_movement"), 160),
            (t("s2_hdr_side"),      60),
        ]
        if is_individual:
            hdr_cols.append(("Rep", 46))
        hdr_cols += [
            (t("s2_hdr_file"), 260),
            ("", 80), ("", 30),
        ]
        for text, w in hdr_cols:
            ctk.CTkLabel(hdr, text=text, width=w,
                         font=ctk.CTkFont(size=11, weight="bold"),
                         anchor="w").pack(side="left", padx=2)

        scroll = ctk.CTkScrollableFrame(card_imp, height=320)
        scroll.pack(fill="x", padx=12, pady=(0, 8))

        from config import MOVEMENT_DEFINITIONS, MOVEMENT_PAIR_GROUPS
        laterality = self._laterality_var.get()

        # Build lookup: first member → (group_label, all_members)
        pair_by_first: dict[str, tuple[str, list[str]]] = {}
        for group_label, members in MOVEMENT_PAIR_GROUPS.items():
            pair_by_first[members[0]] = (group_label, members)
        paired_members: set[str] = {
            mv for members in MOVEMENT_PAIR_GROUPS.values() for mv in members
        }

        rendered_as_pair: set[str] = set()
        selected_set = set(selected_movements)

        for mv_name in selected_movements:
            if mv_name in rendered_as_pair:
                continue

            mv_def = MOVEMENT_DEFINITIONS[mv_name]

            # ── Pair group row (shared file for Thorax + Trunk) ──────────
            if mv_name in pair_by_first:
                group_label, pair_members = pair_by_first[mv_name]
                other_members = [m for m in pair_members if m != mv_name]
                if all(m in selected_set for m in other_members):
                    rendered_as_pair.update(pair_members)
                    reps_range = range(num_reps) if is_individual else range(1)
                    for rep_idx in reps_range:
                        row_frame = ctk.CTkFrame(scroll, fg_color="transparent")
                        row_frame.pack(fill="x", pady=3)
                        main_row = ctk.CTkFrame(row_frame, fg_color="transparent")
                        main_row.pack(fill="x")

                        mv_display = "Thorax/Trunk Ext." if rep_idx == 0 else ""
                        ctk.CTkLabel(main_row, text=mv_display, width=160,
                                     anchor="w", font=ctk.CTkFont(size=11),
                                     ).pack(side="left", padx=2)
                        ctk.CTkLabel(main_row, text=_NO_SIDE if rep_idx == 0 else "",
                                     width=60, anchor="center",
                                     font=ctk.CTkFont(size=11),
                                     ).pack(side="left", padx=2)
                        if is_individual:
                            ctk.CTkLabel(
                                main_row,
                                text=t("s2_rep_label").format(n=rep_idx + 1),
                                width=46, anchor="center",
                                font=ctk.CTkFont(size=11),
                            ).pack(side="left", padx=2)

                        file_lbl = ctk.CTkLabel(
                            main_row, text=t("s2_no_file"), width=260,
                            anchor="w", font=ctk.CTkFont(size=10),
                            text_color="gray",
                        )
                        file_lbl.pack(side="left", padx=2)

                        row_data: dict = {
                            "mv_name":          mv_name,
                            "mv_names":         list(pair_members),
                            "side":             _NO_SIDE,
                            "bilateral":        False,
                            "rep_idx":          rep_idx,
                            "loaded":           False,
                            "filename":         "",
                            "c3d_path":         "",
                            "angle_data":       None,
                            "angle_data_trunk": None,
                            "angle_data_left":  None,
                            "angle_data_right": None,
                            "frame_rate":       None,
                            "events":           None,
                            "file_lbl":         file_lbl,
                            "status_lbl":       None,
                            "offset_var":       ctk.BooleanVar(value=False),
                            "is_individual":    is_individual,
                            "is_pair":          True,
                        }
                        self._import_rows.append(row_data)

                        ctk.CTkButton(
                            main_row, text=t("s2_browse"), width=80,
                            command=lambda rd=row_data: self._browse_c3d(rd),
                        ).pack(side="left", padx=4)

                        status_lbl = ctk.CTkLabel(
                            main_row, text="✗", width=28,
                            font=ctk.CTkFont(size=14), text_color="#E05252",
                        )
                        status_lbl.pack(side="left", padx=2)
                        row_data["status_lbl"] = status_lbl

                        self._restore_row_state(
                            row_data,
                            old_state.get((mv_name, _NO_SIDE, rep_idx)))
                    continue

            # ── Normal single-movement row ────────────────────────────────
            has_prefix = mv_def["has_side_prefix"]
            if not has_prefix:
                row_specs = [(_NO_SIDE, False)]
            elif laterality == "bilateral":
                row_specs = [("L+R", True)]
            else:
                row_specs = [("Left", False), ("Right", False)]

            for side_label, is_bilateral in row_specs:
                reps_range = range(num_reps) if is_individual else range(1)

                for rep_idx in reps_range:
                    row_frame = ctk.CTkFrame(scroll, fg_color="transparent")
                    row_frame.pack(fill="x", pady=3)

                    main_row = ctk.CTkFrame(row_frame, fg_color="transparent")
                    main_row.pack(fill="x")

                    mv_display = _display_name(mv_name) if rep_idx == 0 else ""
                    ctk.CTkLabel(main_row, text=mv_display, width=160,
                                 anchor="w", font=ctk.CTkFont(size=11),
                                 ).pack(side="left", padx=2)

                    side_display = side_label if rep_idx == 0 else ""
                    ctk.CTkLabel(main_row, text=side_display, width=60,
                                 anchor="center", font=ctk.CTkFont(size=11),
                                 ).pack(side="left", padx=2)

                    if is_individual:
                        ctk.CTkLabel(
                            main_row,
                            text=t("s2_rep_label").format(n=rep_idx + 1),
                            width=46,
                            anchor="center", font=ctk.CTkFont(size=11),
                        ).pack(side="left", padx=2)

                    file_lbl = ctk.CTkLabel(
                        main_row, text=t("s2_no_file"), width=260,
                        anchor="w", font=ctk.CTkFont(size=10),
                        text_color="gray",
                    )
                    file_lbl.pack(side="left", padx=2)

                    row_data: dict = {
                        "mv_name":          mv_name,
                        "side":             side_label,
                        "bilateral":        is_bilateral,
                        "rep_idx":          rep_idx,
                        "loaded":           False,
                        "filename":         "",
                        "c3d_path":         "",
                        "angle_data":       None,
                        "angle_data_left":  None,
                        "angle_data_right": None,
                        "frame_rate":       None,
                        "events":           None,
                        "file_lbl":         file_lbl,
                        "status_lbl":       None,
                        "offset_var":       ctk.BooleanVar(value=False),
                        "is_individual":    is_individual,
                    }
                    self._import_rows.append(row_data)

                    ctk.CTkButton(
                        main_row, text=t("s2_browse"), width=80,
                        command=lambda rd=row_data: self._browse_c3d(rd),
                    ).pack(side="left", padx=4)

                    status_lbl = ctk.CTkLabel(
                        main_row, text="✗", width=28,
                        font=ctk.CTkFont(size=14), text_color="#E05252",
                    )
                    status_lbl.pack(side="left", padx=2)
                    row_data["status_lbl"] = status_lbl

                    # Restore state if available
                    self._restore_row_state(
                        row_data,
                        old_state.get((mv_name, side_label, rep_idx)))

        # Navigation buttons
        nav = ctk.CTkFrame(f, fg_color="transparent")
        nav.pack(fill="x", pady=(8, 0))

        ctk.CTkButton(
            nav, text=t("s2_back"), width=200,
            command=self._show_screen_1,
        ).pack(side="left")

        process_cmd = (self._start_individual_processing if is_individual
                       else self._start_segmentation)
        self._process_btn = ctk.CTkButton(
            nav, text=t("s2_process"),
            height=40, width=160,
            font=ctk.CTkFont(size=13, weight="bold"),
            state="disabled",
            command=process_cmd,
        )
        self._process_btn.pack(side="right")
        self._update_process_btn()

    def _restore_row_state(self, row_data: dict, old: dict | None) -> None:
        """Re-apply a previously loaded file's state to newly built widgets."""
        if old is None or not old.get("loaded"):
            return

        for field in ("loaded", "filename", "c3d_path", "angle_data",
                      "angle_data_left", "angle_data_right", "frame_rate", "events"):
            if field in old:
                row_data[field] = old[field]

        # Update widgets to reflect loaded state
        fname = old.get("filename", "…")
        row_data["file_lbl"].configure(text=fname, text_color="white")
        row_data["status_lbl"].configure(text="✓", text_color="#4CAF50")

    def _browse_c3d(self, row_data: dict) -> None:
        path = filedialog.askopenfilename(
            title=f"C3D — {row_data['mv_name']} / {row_data['side']}",
            filetypes=[("C3D files", "*.c3d"), ("All files", "*.*")],
        )
        if not path:
            return

        from data_processing import (
            read_c3d, list_available_angles, compute_trunk_extended_angles,
            compute_thorax_trunk_pair,
        )
        from config import MOVEMENT_DEFINITIONS

        try:
            c3d_data = read_c3d(path)
        except Exception as exc:
            messagebox.showerror(
                t("s2_load_error_title"),
                t("s2_load_error_msg").format(exc=exc),
            )
            return

        fname = os.path.basename(path)

        # ── Paired row (Thorax + Trunk Extended share one file) ───────────
        if row_data.get("is_pair"):
            try:
                result = compute_thorax_trunk_pair(c3d_data)
            except KeyError as exc:
                messagebox.showwarning(
                    t("s2_vars_not_found_title"),
                    t("err_thorax_trunk_vars_missing").format(missing=str(exc)),
                )
                return
            row_data["angle_data"]        = result["thorax_norm_signed"]
            row_data["angle_data_trunk"]  = result["trunk_inclination_z"]
            row_data["loaded"]     = True
            row_data["filename"]   = fname
            row_data["c3d_path"]   = path
            row_data["frame_rate"] = c3d_data["frame_rate"]
            row_data["events"]     = c3d_data.get("events", [])
            row_data["file_lbl"].configure(text=fname, text_color="white")
            row_data["status_lbl"].configure(text="✓", text_color="#4CAF50")
            self._update_process_btn()
            return

        mv_def    = MOVEMENT_DEFINITIONS[row_data["mv_name"]]
        model_outputs = c3d_data["model_outputs"]

        # ── Computed type (marker-based, e.g. Thorax/Trunk Extended Lateral Inclination) ──
        if mv_def.get("type") == "computed":
            try:
                trunk_angles = compute_trunk_extended_angles(c3d_data)
            except (KeyError, ValueError) as exc:
                messagebox.showwarning(
                    t("s2_vars_not_found_title"),
                    t("err_trunk_markers_missing").format(missing=str(exc)),
                )
                return
            row_data["trunk_angles"] = trunk_angles
            row_data["angle_data"]   = trunk_angles[mv_def["primary_key"]]

        # ── Standard SULM model-output type ───────────────────────────────
        else:
            has_prefix = mv_def["has_side_prefix"]
            component  = mv_def["component"]
            sulm_var   = mv_def["sulm_variable"]

            if row_data["bilateral"]:
                var_left  = f"Left{sulm_var}"
                var_right = f"Right{sulm_var}"
                missing = [v for v in (var_left, var_right)
                           if v not in model_outputs]
                if missing:
                    available = list_available_angles(c3d_data)
                    messagebox.showwarning(
                        t("s2_vars_not_found_title"),
                        t("s2_vars_not_found_msg").format(
                            missing="\n  ".join(missing),
                            available="\n  ".join(available),
                        ),
                    )
                    return
                row_data["angle_data_left"]  = (
                    model_outputs[var_left][component, :].astype(float))
                row_data["angle_data_right"] = (
                    model_outputs[var_right][component, :].astype(float))
                row_data["angle_data"] = row_data["angle_data_left"]

            else:
                side     = row_data["side"]
                full_var = (f"{side}{sulm_var}"
                            if has_prefix and side != _NO_SIDE else sulm_var)
                if full_var not in model_outputs:
                    available = list_available_angles(c3d_data)
                    messagebox.showwarning(
                        t("s2_var_not_found_title"),
                        t("s2_var_not_found_msg").format(
                            label=full_var,
                            available="\n  ".join(available),
                        ),
                    )
                    return
                row_data["angle_data"] = (
                    model_outputs[full_var][component, :].astype(float))

        row_data["loaded"]     = True
        row_data["filename"]   = fname
        row_data["c3d_path"]   = path
        row_data["frame_rate"] = c3d_data["frame_rate"]
        row_data["events"]     = c3d_data.get("events", [])

        row_data["file_lbl"].configure(text=fname, text_color="white")
        row_data["status_lbl"].configure(text="✓", text_color="#4CAF50")

        self._update_process_btn()

    def _update_process_btn(self) -> None:
        all_loaded = all(r["loaded"] for r in self._import_rows)
        if self._process_btn is not None:
            self._process_btn.configure(
                state="normal" if all_loaded else "disabled")

    # ══════════════════════════════════════════════════════════════════════
    #  Screen 3 — Segmentation (modal windows per movement+side)
    # ══════════════════════════════════════════════════════════════════════

    @staticmethod
    def _expand_pair_rows(rows: list[dict]) -> list[dict]:
        """Expand paired file rows into two virtual single-movement rows."""
        from config import MOVEMENT_DEFINITIONS
        out = []
        for r in rows:
            if r.get("is_pair"):
                for mv_name in r["mv_names"]:
                    data_key = MOVEMENT_DEFINITIONS[mv_name].get(
                        "pair_data_key", "angle_data")
                    out.append({**r, "mv_name": mv_name,
                                "angle_data": r.get(data_key),
                                "is_pair": False})
            else:
                out.append(r)
        return out

    def _start_segmentation(self) -> None:
        from segmentation import C3DSegmentationWindow
        from data_processing import apply_offset, compute_extended_stats_array
        from config import MOVEMENT_DEFINITIONS

        rows_to_process = self._expand_pair_rows(self._import_rows)

        # Build a flat list of all items to segment upfront so we can index
        # into it freely (needed for back-navigation).
        all_items: list[dict] = []
        for row_data in rows_to_process:
            mv_name     = row_data["mv_name"]
            frame_rate  = row_data["frame_rate"]
            events      = row_data["events"] or []
            offset_on   = row_data["offset_var"].get()
            mv_def      = MOVEMENT_DEFINITIONS[mv_name]
            is_computed = mv_def.get("type") == "computed"

            if row_data["bilateral"]:
                try:
                    off_l = float(row_data["offset_left_var"].get()) if offset_on else 0.0
                except (ValueError, KeyError):
                    off_l = 0.0
                try:
                    off_r = float(row_data["offset_right_var"].get()) if offset_on else 0.0
                except (ValueError, KeyError):
                    off_r = 0.0
                sides_to_seg: list[tuple] = [
                    ("Left",
                     apply_offset(row_data["angle_data_left"], off_l) if offset_on
                     else row_data["angle_data_left"],
                     off_l if offset_on else None,
                     None),
                    ("Right",
                     apply_offset(row_data["angle_data_right"], off_r) if offset_on
                     else row_data["angle_data_right"],
                     off_r if offset_on else None,
                     None),
                ]
            else:
                try:
                    off = float(row_data["offset_entry_var"].get()) if offset_on else 0.0
                except (ValueError, KeyError):
                    off = 0.0
                ang         = apply_offset(row_data["angle_data"], off) if offset_on else row_data["angle_data"]
                trunk_angles = row_data.get("trunk_angles") if is_computed else None
                sides_to_seg = [(row_data["side"], ang, off if offset_on else None, trunk_angles)]

            for side, angle_arr, offset_val, trunk_ang in sides_to_seg:
                all_items.append({
                    "mv_name":     mv_name,
                    "side":        side,
                    "angle_arr":   angle_arr,
                    "offset_val":  offset_val,
                    "trunk_ang":   trunk_ang,
                    "frame_rate":  frame_rate,
                    "events":      events,
                    "is_computed": is_computed,
                    "c3d_path":    row_data.get("c3d_path", ""),
                })

        last_direction = "peak_to_valley"
        i = 0
        while i < len(all_items):
            item       = all_items[i]
            mv_name    = item["mv_name"]
            side       = item["side"]
            angle_arr  = item["angle_arr"]
            offset_val = item["offset_val"]
            trunk_ang  = item["trunk_ang"]
            frame_rate = item["frame_rate"]
            events     = item["events"]
            is_computed = item["is_computed"]
            c3d_path   = item["c3d_path"]

            title = f"{mv_name} — {side}" if side != _NO_SIDE else mv_name

            if is_computed and trunk_ang is not None:
                win = _TrunkSegmentationWindow(
                    self,
                    movement_name=title,
                    angle_data=angle_arr,
                    trunk_angles=trunk_ang,
                    frame_rate=frame_rate,
                    events=events,
                )
            else:
                win = C3DSegmentationWindow(
                    self,
                    movement_name=title,
                    angle_data=angle_arr,
                    frame_rate=frame_rate,
                    events=events,
                    show_back=(i > 0),
                    initial_direction=last_direction,
                )
            self.wait_window(win)

            if win.go_back:
                i -= 1
                prev = all_items[i]
                self._processed.pop((prev["mv_name"], prev["side"]), None)
                continue

            last_direction = win.last_direction

            if win.result is not None:
                segs = win.result.get("segments", [])
                self._processed[(mv_name, side)] = {
                    "movement":   mv_name,
                    "side":       side,
                    "angle_data": angle_arr,
                    "frame_rate": frame_rate,
                    "offset":     offset_val,
                    "c3d_path":   c3d_path,
                    "extended":   compute_extended_stats_array(angle_arr, segs),
                    **win.result,
                }
            else:
                logger.info("Segmentation cancelled for '%s' — %s.", mv_name, side)

            i += 1

        if not self._processed:
            messagebox.showinfo(
                t("s3_no_results_title"),
                t("s3_no_results_msg"),
            )
            return

        self._show_screen_4()

    # ══════════════════════════════════════════════════════════════════════
    #  Individual mode processing
    # ══════════════════════════════════════════════════════════════════════

    def _start_individual_processing(self) -> None:
        from individual_review import IndividualReviewWindow
        from data_processing import apply_offset

        # Expand pair rows before grouping
        expanded = self._expand_pair_rows(self._import_rows)

        # Group rows by (mv_name, side)
        groups: dict[tuple[str, str], list[dict]] = {}
        for r in expanded:
            key = (r["mv_name"], r["side"])
            groups.setdefault(key, []).append(r)

        for (mv_name, side), reps in groups.items():
            # Sort by rep_idx to keep order stable
            reps = sorted(reps, key=lambda r: r["rep_idx"])

            offset_on = reps[0]["offset_var"].get()

            if reps[0]["bilateral"]:
                # Bilateral Individual: process Left and Right separately
                for actual_side, data_key, off_key in [
                    ("Left",  "angle_data_left",  "offset_left_var"),
                    ("Right", "angle_data_right", "offset_right_var"),
                ]:
                    try:
                        off = (float(reps[0][off_key].get())
                               if offset_on else 0.0)
                    except (ValueError, KeyError):
                        off = 0.0

                    curves = [r[data_key] for r in reps if r[data_key] is not None]
                    if not curves:
                        continue

                    title = f"{mv_name} — {actual_side}"
                    win = IndividualReviewWindow(
                        self, title, curves,
                        offset_val=off if offset_on else None,
                    )
                    self.wait_window(win)

                    if win.result is not None:
                        self._processed[(mv_name, actual_side)] = {
                            "movement":   mv_name,
                            "side":       actual_side,
                            "angle_data": win.result["angle_data"],
                            "frame_rate": reps[0]["frame_rate"],
                            "offset":     win.result["offset"],
                            "c3d_path":   reps[0].get("c3d_path", ""),
                            "extended":   win.result["extended"],
                            "segments":   [],
                        }
            else:
                try:
                    off = (float(reps[0]["offset_entry_var"].get())
                           if offset_on else 0.0)
                except (ValueError, KeyError):
                    off = 0.0

                curves = [r["angle_data"] for r in reps
                          if r["angle_data"] is not None]
                if not curves:
                    continue

                title = (f"{mv_name} — {side}"
                         if side != _NO_SIDE else mv_name)
                win = IndividualReviewWindow(
                    self, title, curves,
                    offset_val=off if offset_on else None,
                )
                self.wait_window(win)

                if win.result is not None:
                    self._processed[(mv_name, side)] = {
                        "movement":   mv_name,
                        "side":       side,
                        "angle_data": win.result["angle_data"],
                        "frame_rate": reps[0]["frame_rate"],
                        "offset":     win.result["offset"],
                        "c3d_path":   reps[0].get("c3d_path", ""),
                        "extended":   win.result["extended"],
                        "segments":   [],
                    }

        if not self._processed:
            messagebox.showinfo(
                t("s3_no_results_title"),
                t("s3_no_results_msg"),
            )
            return

        self._show_screen_4()

    # ══════════════════════════════════════════════════════════════════════
    #  Screen 4 — Summary & Export
    # ══════════════════════════════════════════════════════════════════════

    def _show_screen_4(self) -> None:
        self._current_screen = 4
        self._clear_container()
        self._charts = {}
        # _layout_vertical preserved across toggles and re-renders

        f = ctk.CTkFrame(self._container, fg_color="transparent")
        f.pack(fill="both", expand=True)

        # ── Top bar: layout toggle ─────────────────────────────────────────
        top_bar = ctk.CTkFrame(f, fg_color="transparent")
        top_bar.pack(fill="x", pady=(0, 4))

        toggle_text = (t("s4_layout_horizontal") if self._layout_vertical
                       else t("s4_layout_vertical"))
        ctk.CTkButton(
            top_bar, text=toggle_text, width=190,
            command=self._toggle_layout,
        ).pack(side="right")

        # ── Scrollable area — one section per movement ─────────────────────
        scroll = ctk.CTkScrollableFrame(f)
        scroll.pack(fill="both", expand=True)

        _is_dark = ctk.get_appearance_mode() == "Dark"
        _bg_alt = "#2A2D2E" if _is_dark else "#EBEBEB"

        from config import MOVEMENT_PAIR_GROUPS
        groups: dict[str, list[tuple[str, dict]]] = {}
        for (mv_name, side), data in self._processed.items():
            groups.setdefault(mv_name, []).append((side, data))

        from config import MOVEMENT_DEFINITIONS
        _def_order = {name: i for i, name in enumerate(MOVEMENT_DEFINITIONS)}
        _pair_members: set[str] = {mv for members in MOVEMENT_PAIR_GROUPS.values() for mv in members}
        _pair_sub: dict[str, int] = {}
        _sub = 0
        for members in MOVEMENT_PAIR_GROUPS.values():
            for mv in members:
                _pair_sub[mv] = _sub
                _sub += 1

        def _sort_key(item):
            name = item[0]
            if name in _pair_members:
                return (999 + _pair_sub.get(name, 0), name)
            return (_def_order.get(name, 500), name)

        for mv_name, sides_data in sorted(groups.items(), key=_sort_key):
            self._build_movement_section(scroll, mv_name, sides_data, _bg_alt)

        # ── Fixed bottom action bar ────────────────────────────────────────
        btn_row = ctk.CTkFrame(f, fg_color="transparent")
        btn_row.pack(fill="x", pady=(6, 0))

        ctk.CTkButton(
            btn_row, text=t("s4_export_xlsx"), width=130,
            command=self._export_xlsx_dialog,
        ).pack(side="left")
        ctk.CTkButton(
            btn_row, text=t("s4_export_csv"), width=130,
            command=self._export_csv_dialog,
        ).pack(side="left", padx=(6, 0))
        ctk.CTkButton(
            btn_row, text=t("s4_new_analysis"), width=160,
            font=ctk.CTkFont(weight="bold"),
            command=self._new_analysis,
        ).pack(side="right")
        ctk.CTkButton(
            btn_row, text=t("s4_generate_report"), width=160,
            command=self._generate_report,
        ).pack(side="right", padx=4)

    def _toggle_layout(self) -> None:
        self._layout_vertical = not self._layout_vertical
        self._show_screen_4()

    @staticmethod
    def _metric_labels_for(mv_name: str) -> list[tuple[str, str]]:
        """Return [(display_label, metric_key), ...] for the 3 metrics of a movement."""
        from config import MOVEMENT_DEFINITIONS
        mv_def = MOVEMENT_DEFINITIONS.get(mv_name, {})
        if "peak_label_key" in mv_def:
            peak_label   = f"{t(mv_def['peak_label_key'])} ({t('lbl_maximum')})"
            valley_label = f"{t(mv_def['valley_label_key'])} ({t('lbl_minimum')})"
        else:
            peak_label   = t("s4_metric_peak")
            valley_label = t("s4_metric_valley")
        return [
            (t("s4_metric_rom"), "rom"),
            (peak_label,         "peak"),
            (valley_label,       "valley"),
        ]

    def _build_movement_section(
        self,
        parent,
        mv_name: str,
        sides_data: list[tuple[str, dict]],
        bg_alt: str,
    ) -> None:
        section = ctk.CTkFrame(parent)
        section.pack(fill="x", padx=4, pady=(0, 10))

        ctk.CTkLabel(
            section, text=_display_name(mv_name),
            font=ctk.CTkFont(size=13, weight="bold"),
        ).pack(anchor="w", padx=12, pady=(8, 4))

        content = ctk.CTkFrame(section, fg_color="transparent")
        content.pack(fill="x", padx=8, pady=(0, 8))

        if self._layout_vertical:
            chart_frame = ctk.CTkFrame(content, fg_color="transparent")
            chart_frame.pack(fill="x", pady=(0, 4))
            self._build_movement_chart(chart_frame, sides_data, mv_name)

            tbl_frame = ctk.CTkFrame(content, fg_color="transparent")
            tbl_frame.pack(fill="x")
            self._build_movement_table(tbl_frame, sides_data, bg_alt, mv_name)
        else:
            tbl_frame = ctk.CTkFrame(content, fg_color="transparent")
            tbl_frame.pack(side="left", fill="both", expand=True, padx=(0, 4))
            self._build_movement_table(tbl_frame, sides_data, bg_alt, mv_name)

            chart_frame = ctk.CTkFrame(content, fg_color="transparent")
            chart_frame.pack(side="left", fill="both", expand=True)
            self._build_movement_chart(chart_frame, sides_data, mv_name)

    def _build_trunk_pair_chart(
        self,
        sides_data: list[tuple[str, dict]],
        mv_name: str,
    ):
        """Raincloud chart for trunk pair movements: 3 subplots ROM/Right/Left, independent Y axes."""
        from matplotlib.figure import Figure
        from matplotlib.patches import Rectangle
        import numpy as np
        try:
            from scipy.stats import gaussian_kde
            _HAS_SCIPY = True
        except ImportError:
            _HAS_SCIPY = False

        _ml = self._metric_labels_for(mv_name)
        METRICS = [
            (_ml[0][1], _ml[0][0], "#3B82F6", "#1D4ED8"),
            (_ml[1][1], _ml[1][0], "#22C55E", "#15803D"),
            (_ml[2][1], _ml[2][0], "#EF4444", "#B91C1C"),
        ]
        _V_AMP, _BW, _SC_OFF, _SC_JIT = 0.30, 0.035, 0.10, 0.03

        extended = {}
        for _, data in sides_data:
            extended = data.get("extended", {})
            break

        fig = Figure(figsize=(9, 4.5))
        axes = fig.subplots(1, 3)
        fig.subplots_adjust(wspace=0.35, top=0.82)
        fig.suptitle(mv_name, fontsize=13, fontweight="bold")
        rng = np.random.default_rng(42)

        for m_idx, (metric_key, metric_label, color, color_dark) in enumerate(METRICS):
            cur_ax = axes[m_idx]
            cur_ax.set_title(metric_label, fontsize=10, fontweight="bold")
            raw  = extended.get(metric_key, {}).get("values", [])
            vals = np.array(
                [v for v in raw if v is not None and not np.isnan(float(v))],
                dtype=float,
            )
            x_c = 0.0

            if vals.size >= 1:
                if vals.size >= 2 and _HAS_SCIPY:
                    kde     = gaussian_kde(vals, bw_method="scott")
                    spread  = max(float(vals.std()) * 0.5, 0.5)
                    y_grid  = np.linspace(vals.min() - spread, vals.max() + spread, 300)
                    density = kde(y_grid)
                    peak    = density.max()
                    if peak > 0:
                        kde_x = x_c - (density / peak) * _V_AMP
                        cur_ax.fill_betweenx(y_grid, kde_x, x_c,
                                             color=color, alpha=0.45, zorder=2)
                        cur_ax.plot(kde_x, y_grid, color=color, alpha=0.75,
                                    linewidth=1.0, zorder=3)

                q1, q2, q3 = np.percentile(vals, [25, 50, 75])
                iqr      = q3 - q1
                fence_lo = q1 - 1.5 * iqr
                fence_hi = q3 + 1.5 * iqr
                w_lo = vals[vals >= fence_lo].min() if np.any(vals >= fence_lo) else q1
                w_hi = vals[vals <= fence_hi].max() if np.any(vals <= fence_hi) else q3
                cur_ax.plot([x_c, x_c], [w_lo, q1], color=color, lw=1.2, zorder=4)
                cur_ax.plot([x_c, x_c], [q3, w_hi], color=color, lw=1.2, zorder=4)
                for wy in (w_lo, w_hi):
                    cur_ax.plot([x_c - _BW, x_c + _BW], [wy, wy],
                                color=color, lw=1.2, zorder=4)
                cur_ax.add_patch(Rectangle(
                    (x_c - _BW, q1), 2 * _BW, q3 - q1,
                    facecolor="white", edgecolor=color, linewidth=1.5, zorder=5,
                ))
                cur_ax.plot([x_c - _BW, x_c + _BW], [q2, q2],
                            color=color, lw=2.0, zorder=6)
                mean_val = float(vals.mean())
                cur_ax.text(x_c, w_hi, f"{mean_val:.1f}°",
                            ha="center", va="bottom", fontsize=10,
                            fontweight="bold", color=color_dark, zorder=7)
                outliers = vals[(vals < w_lo) | (vals > w_hi)]
                if outliers.size:
                    cur_ax.scatter(np.full(outliers.size, x_c), outliers,
                                   color=color, s=25, marker="D", alpha=0.85, zorder=7)
                jitter = rng.uniform(-_SC_JIT, _SC_JIT, size=vals.size)
                cur_ax.scatter(x_c + _SC_OFF + jitter, vals,
                               color=color, s=35, alpha=0.85, zorder=3)

            cur_ax.set_xticks([x_c])
            cur_ax.set_xticklabels([metric_label], fontsize=9)
            cur_ax.set_xlim(-0.65, 0.65)
            cur_ax.tick_params(axis="x", length=0)
            if m_idx == 0:
                cur_ax.set_ylabel(t("ylabel_degrees"), fontsize=9)
            cur_ax.margins(y=0.2)
            cur_ax.spines["top"].set_visible(False)
            cur_ax.spines["right"].set_visible(False)
            cur_ax.yaxis.grid(True, linestyle="--", alpha=0.4, zorder=0)

        from matplotlib.ticker import MaxNLocator
        for ax_ in axes:
            ax_.yaxis.set_major_locator(MaxNLocator(nbins=5, integer=False))

        return fig

    def _build_movement_chart(
        self,
        parent,
        sides_data: list[tuple[str, dict]],
        mv_name: str = "",
    ) -> None:
        from config import MOVEMENT_PAIR_GROUPS
        _pair_members = {mv for members in MOVEMENT_PAIR_GROUPS.values()
                         for mv in members}

        if mv_name in _pair_members:
            fig = self._build_trunk_pair_chart(sides_data, mv_name)
        else:
            from plotting import plot_rom_raincloud
            ml = self._metric_labels_for(mv_name)
            fig = plot_rom_raincloud(
                {mv_name: sides_data},
                ylabel=t("ylabel_degrees"),
                metric_labels=[(key, lbl) for lbl, key in ml],
            )

        self._charts[mv_name] = fig
        try:
            canvas = FigureCanvasTkAgg(fig, master=parent)
            canvas.draw()
            canvas.get_tk_widget().pack(fill="both", expand=True)
            ctk.CTkButton(
                parent, text=t("export_chart"), width=130,
                command=lambda f=fig, n=mv_name: self._export_chart(f, n),
            ).pack(anchor="e", padx=4, pady=(2, 0))
        except Exception as exc:
            ctk.CTkLabel(parent, text=t("s4_chart_unavailable").format(exc=exc),
                         text_color="gray").pack(padx=8, pady=4)

    def _build_movement_table(
        self,
        parent,
        sides_data: list[tuple[str, dict]],
        bg_alt: str,
        mv_name: str = "",
    ) -> None:
        headers = [
            t("s4_hdr_side"), t("s4_hdr_metric"),
            t("s4_hdr_n"), t("s4_hdr_mean"), t("s4_hdr_sd"),
            t("s4_hdr_min"), t("s4_hdr_max"),
        ]
        col_w = [58, 65, 32, 68, 56, 56, 56]

        scroll = ctk.CTkScrollableFrame(parent, height=130)
        scroll.pack(fill="both", expand=True, padx=4, pady=(0, 4))

        for c, (h, w) in enumerate(zip(headers, col_w)):
            ctk.CTkLabel(scroll, text=h, width=w,
                         font=ctk.CTkFont(size=11, weight="bold"),
                         anchor="w").grid(row=0, column=c, padx=2, pady=2,
                                          sticky="w")

        metric_labels = self._metric_labels_for(mv_name)

        offset_notes: list[str] = []
        row_idx = 1

        for group_idx, (side, data) in enumerate(sides_data):
            extended   = data.get("extended", {})
            offset_val = data.get("offset")

            if offset_val is not None:
                offset_notes.append(
                    f"({side}): "
                    + t("s4_offset_subtracted").format(offset=offset_val)
                )

            bg = bg_alt if group_idx % 2 == 0 else "transparent"

            for sub_idx, (metric_label, metric_key) in enumerate(metric_labels):
                stats  = extended.get(metric_key, {})
                values = stats.get("values", [])
                mean_v = stats.get("mean", float("nan"))
                sd_v   = stats.get("sd",   0.0)
                min_v  = stats.get("min",  float("nan"))
                max_v  = stats.get("max",  float("nan"))

                row_vals = [
                    side if sub_idx == 0 else "",
                    metric_label,
                    str(len(values)),
                    f"{mean_v:.1f}" if not math.isnan(mean_v) else "—",
                    f"{sd_v:.1f}",
                    f"{min_v:.1f}"  if not math.isnan(min_v)  else "—",
                    f"{max_v:.1f}"  if not math.isnan(max_v)  else "—",
                ]

                for c, (val, w) in enumerate(zip(row_vals, col_w)):
                    ctk.CTkLabel(scroll, text=val, width=w,
                                 font=ctk.CTkFont(size=11), anchor="w",
                                 fg_color=bg).grid(row=row_idx, column=c,
                                                   padx=2, pady=1, sticky="w")
                row_idx += 1

        if offset_notes:
            note = (t("s4_offset_note_prefix")
                    + "; ".join(offset_notes) + ".")
            ctk.CTkLabel(parent, text=note,
                         font=ctk.CTkFont(size=9), text_color="gray",
                         anchor="w", wraplength=380,
                         ).pack(fill="x", padx=4, pady=(0, 2))

    # ── Export ─────────────────────────────────────────────────────────────

    def _export_csv(self) -> None:
        if not self._processed:
            return

        out_path = filedialog.asksaveasfilename(
            title=t("s4_csv_dialog_title"),
            defaultextension=".csv",
            filetypes=[("CSV", "*.csv")],
            initialfile="ROM_Summary.csv",
        )
        if not out_path:
            return

        rows = []
        for (mv_name, side), data in self._processed.items():
            extended   = data.get("extended", {})
            offset_val = data.get("offset")

            for metric_label, metric_key in self._metric_labels_for(mv_name):
                stats  = extended.get(metric_key, {})
                values = stats.get("values", [])
                m  = stats.get("mean", float("nan"))
                s  = stats.get("sd",   0.0)
                mn = stats.get("min",  float("nan"))
                mx = stats.get("max",  float("nan"))

                rows.append({
                    "Movement":   mv_name,
                    "Side":       side,
                    "Metric":     metric_label,
                    "N_reps":     len(values),
                    "Mean_deg":   round(m,  2) if not math.isnan(m)  else None,
                    "SD_deg":     round(s,  2),
                    "Min_deg":    round(mn, 2) if not math.isnan(mn) else None,
                    "Max_deg":    round(mx, 2) if not math.isnan(mx) else None,
                    "Values_deg": ";".join(
                        f"{v:.2f}" if not math.isnan(v) else "nan"
                        for v in values
                    ),
                    "Offset_deg": (f"{offset_val:.2f}"
                                   if offset_val is not None else ""),
                })

        pd.DataFrame(rows).to_csv(out_path, index=False)

        msg = t("s4_csv_saved_msg").format(path=out_path)
        offset_entries = [
            (mv, sd, d["offset"])
            for (mv, sd), d in self._processed.items()
            if d.get("offset") is not None
        ]
        if offset_entries:
            notes = [
                t("s4_csv_offset_entry").format(mv=mv, sd=sd, off=off)
                for mv, sd, off in offset_entries
            ]
            msg += t("s4_csv_offset_applied").format(
                notes="\n".join(notes))
        messagebox.showinfo(t("s4_csv_saved_title"), msg)

    def _export_csv_dialog(self) -> None:
        if not self._processed:
            return

        win = ctk.CTkToplevel(self)
        win.title(t("s4_csv_dlg_win_title"))
        win.geometry("340x300")
        win.resizable(False, False)
        win.grab_set()
        win.lift()
        win.focus_force()

        ctk.CTkLabel(
            win, text=t("s4_xlsx_select_sheets"),
            font=ctk.CTkFont(size=13, weight="bold"),
        ).pack(padx=20, pady=(16, 8))

        var_summary = ctk.BooleanVar(value=True)
        var_rep     = ctk.BooleanVar(value=True)
        var_raw     = ctk.BooleanVar(value=True)

        ctk.CTkCheckBox(
            win, text=t("s4_xlsx_sheet_summary"), variable=var_summary,
        ).pack(anchor="w", padx=32, pady=4)
        ctk.CTkCheckBox(
            win, text=t("s4_xlsx_sheet_rep_detail"), variable=var_rep,
        ).pack(anchor="w", padx=32, pady=4)
        ctk.CTkCheckBox(
            win, text=t("s4_xlsx_sheet_raw_data"), variable=var_raw,
        ).pack(anchor="w", padx=32, pady=4)

        def _do_export():
            selected = []
            if var_summary.get():
                selected.append("summary")
            if var_rep.get():
                selected.append("rep_detail")
            if var_raw.get():
                selected.append("raw_data")

            if not selected:
                messagebox.showwarning(
                    t("s4_xlsx_no_sheets_title"),
                    t("s4_xlsx_no_sheets_msg"),
                    parent=win,
                )
                return

            path = filedialog.asksaveasfilename(
                parent=win,
                title=t("s4_csv_dlg_file_title"),
                defaultextension=".csv",
                filetypes=[("CSV", "*.csv")],
                initialfile="ROM_Summary.csv",
            )
            if not path:
                return
            win.destroy()
            self._build_and_save_csv(path, selected)

        btn_frame = ctk.CTkFrame(win, fg_color="transparent")
        btn_frame.pack(fill="x", padx=20, pady=(12, 16))
        ctk.CTkButton(
            btn_frame, text=t("s4_xlsx_export_btn"), command=_do_export,
        ).pack(side="right")

    def _build_and_save_csv(self, path: str, sections: list) -> None:
        import os
        stem, _ = os.path.splitext(path)

        builders = {
            "summary":    self._csv_build_summary,
            "rep_detail": self._csv_build_rep_detail,
            "raw_data":   self._csv_build_raw_data,
        }

        saved = []
        for section in sections:
            out = path if len(sections) == 1 else f"{stem}_{section}.csv"
            builders[section]().to_csv(out, index=False)
            saved.append(out)

        messagebox.showinfo(
            t("s4_csv_dlg_saved_title"),
            t("s4_csv_dlg_saved_msg").format(paths="\n".join(saved)),
        )

    def _csv_build_summary(self) -> "pd.DataFrame":
        rows = []
        for (mv_name, side), data in self._processed.items():
            extended = data.get("extended", {})
            mv_label = mv_name
            for metric_label, metric_key in self._metric_labels_for(mv_name):
                stats  = extended.get(metric_key, {})
                values = stats.get("values", [])
                m  = stats.get("mean", float("nan"))
                s  = stats.get("sd",   0.0)
                mn = stats.get("min",  float("nan"))
                mx = stats.get("max",  float("nan"))
                rows.append({
                    "Movement": mv_label,
                    "Side":     side,
                    "Metric":   metric_label,
                    "N_reps":   len(values),
                    "Mean_deg": round(m,  2) if not math.isnan(m)  else None,
                    "SD_deg":   round(s,  2),
                    "Min_deg":  round(mn, 2) if not math.isnan(mn) else None,
                    "Max_deg":  round(mx, 2) if not math.isnan(mx) else None,
                })
        return pd.DataFrame(rows)

    def _csv_build_rep_detail(self) -> "pd.DataFrame":
        import numpy as _np
        rows = []
        for (mv_name, side), data in self._processed.items():
            segments  = data.get("segments", [])
            angle_arr = data.get("angle_data")
            extended  = data.get("extended", {})
            frame_rate = float(data.get("frame_rate") or 100.0)

            if segments and isinstance(angle_arr, _np.ndarray):
                for i, (s, e) in enumerate(segments):
                    chunk = angle_arr[s : e + 1]
                    valid = chunk[~_np.isnan(chunk)]
                    if valid.size == 0:
                        rows.append({
                            "Movement": mv_name, "Side": side, "Rep": i + 1,
                            "Peak_frame": None, "Peak_time_s": None, "Peak_deg": None,
                            "Valley_frame": None, "Valley_time_s": None, "Valley_deg": None,
                            "ROM_deg": None,
                        })
                    else:
                        pk_fr  = s + int(_np.nanargmax(chunk))
                        vl_fr  = s + int(_np.nanargmin(chunk))
                        pk_val = float(_np.nanmax(chunk))
                        vl_val = float(_np.nanmin(chunk))
                        rows.append({
                            "Movement":     mv_name,
                            "Side":         side,
                            "Rep":          i + 1,
                            "Peak_frame":   pk_fr,
                            "Peak_time_s":  round(pk_fr / frame_rate, 4),
                            "Peak_deg":     round(pk_val, 2),
                            "Valley_frame": vl_fr,
                            "Valley_time_s": round(vl_fr / frame_rate, 4),
                            "Valley_deg":   round(vl_val, 2),
                            "ROM_deg":      round(pk_val - vl_val, 2),
                        })
            else:
                peak_vals   = extended.get("peak",   {}).get("values", [])
                valley_vals = extended.get("valley", {}).get("values", [])
                rom_vals    = extended.get("rom",    {}).get("values", [])
                n = max(len(peak_vals), len(valley_vals), len(rom_vals), 0)
                for i in range(n):
                    pk = peak_vals[i]   if i < len(peak_vals)   else float("nan")
                    vl = valley_vals[i] if i < len(valley_vals) else float("nan")
                    rm = rom_vals[i]    if i < len(rom_vals)    else float("nan")
                    rows.append({
                        "Movement":     mv_name,
                        "Side":         side,
                        "Rep":          i + 1,
                        "Peak_frame":   None,
                        "Peak_time_s":  None,
                        "Peak_deg":     round(pk, 2) if not math.isnan(pk) else None,
                        "Valley_frame": None,
                        "Valley_time_s": None,
                        "Valley_deg":   round(vl, 2) if not math.isnan(vl) else None,
                        "ROM_deg":      round(rm, 2) if not math.isnan(rm) else None,
                    })
        return pd.DataFrame(rows)

    def _csv_build_raw_data(self) -> "pd.DataFrame":
        import numpy as _np
        rows = []
        for (mv_name, side), data in self._processed.items():
            angle_arr  = data.get("angle_data")
            frame_rate = float(data.get("frame_rate") or 100.0)
            if not isinstance(angle_arr, _np.ndarray):
                continue
            for fr, val in enumerate(angle_arr):
                rows.append({
                    "Movement": mv_name,
                    "Side":     side,
                    "Frame":    fr,
                    "Time_s":   round(fr / frame_rate, 4),
                    "Angle_deg": None if math.isnan(float(val)) else round(float(val), 4),
                })
        return pd.DataFrame(rows)

    def _export_xlsx_dialog(self) -> None:
        if not self._processed:
            return

        win = ctk.CTkToplevel(self)
        win.title(t("s4_xlsx_win_title"))
        win.geometry("340x360")
        win.resizable(False, False)
        win.grab_set()
        win.lift()
        win.focus_force()

        ctk.CTkLabel(
            win, text=t("s4_xlsx_select_sheets"),
            font=ctk.CTkFont(size=13, weight="bold"),
        ).pack(padx=20, pady=(16, 8))

        var_summary  = ctk.BooleanVar(value=True)
        var_rep      = ctk.BooleanVar(value=True)
        var_raw      = ctk.BooleanVar(value=True)

        ctk.CTkCheckBox(
            win, text=t("s4_xlsx_sheet_summary"), variable=var_summary,
        ).pack(anchor="w", padx=32, pady=4)
        ctk.CTkCheckBox(
            win, text=t("s4_xlsx_sheet_rep_detail"), variable=var_rep,
        ).pack(anchor="w", padx=32, pady=4)
        ctk.CTkCheckBox(
            win, text=t("s4_xlsx_sheet_raw_data"), variable=var_raw,
        ).pack(anchor="w", padx=32, pady=4)

        def _do_export():
            selected = []
            if var_summary.get():
                selected.append("summary")
            if var_rep.get():
                selected.append("rep_detail")
            if var_raw.get():
                selected.append("raw_data")

            if not selected:
                messagebox.showwarning(
                    t("s4_xlsx_no_sheets_title"),
                    t("s4_xlsx_no_sheets_msg"),
                    parent=win,
                )
                return
            path = filedialog.asksaveasfilename(
                parent=win,
                title=t("s4_xlsx_dialog_title"),
                defaultextension=".xlsx",
                filetypes=[("Excel", "*.xlsx")],
                initialfile="ROM_Summary.xlsx",
            )
            if not path:
                return
            win.destroy()
            self._build_and_save_xlsx(path, selected)

        btn_frame = ctk.CTkFrame(win, fg_color="transparent")
        btn_frame.pack(fill="x", padx=20, pady=(12, 16))
        ctk.CTkButton(
            btn_frame, text=t("s4_xlsx_export_btn"), command=_do_export,
        ).pack(side="right")

    def _build_and_save_xlsx(
        self,
        path: str,
        sheets: list,
    ) -> None:
        import openpyxl

        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        groups: dict = {}
        for (mv_name, side), data in self._processed.items():
            groups.setdefault(mv_name, []).append((side, data))

        if "summary" in sheets:
            ws = wb.create_sheet(t("s4_xlsx_sheet_summary"))
            self._xl_write_summary(ws, groups)

        if "rep_detail" in sheets:
            ws = wb.create_sheet(t("s4_xlsx_sheet_rep_detail"))
            self._xl_write_rep_detail(ws, groups)

        if "raw_data" in sheets:
            ws = wb.create_sheet(t("s4_xlsx_sheet_raw_data"))
            self._xl_write_raw_data(ws, groups)

        wb.save(path)
        messagebox.showinfo(
            t("s4_xlsx_saved_title"),
            t("s4_xlsx_saved_msg").format(path=path),
        )

    def _xl_write_summary(self, ws, groups: dict) -> None:
        import io
        import math as _math
        from openpyxl.styles import Font
        from openpyxl.drawing.image import Image as XLImage

        bold = Font(bold=True)
        headers = [
            t("s4_hdr_movement"), t("s4_hdr_side"), t("s4_hdr_metric"),
            t("s4_hdr_n"), t("s4_hdr_mean"), t("s4_hdr_sd"),
            t("s4_hdr_min"), t("s4_hdr_max"),
        ]
        for c, h in enumerate(headers, 1):
            ws.cell(row=1, column=c, value=h).font = bold

        row = 2
        for mv_name, sides_data in groups.items():
            is_bilateral = len({d.get("c3d_path", "") for _, d in sides_data}) == 1
            mv_label = mv_name + (" (Bilateral)" if is_bilateral else " (Unilateral)")
            mv_start = row
            for side, data in sides_data:
                extended = data.get("extended", {})
                for metric_label, metric_key in self._metric_labels_for(mv_name):
                    stats  = extended.get(metric_key, {})
                    values = stats.get("values", [])
                    mean_v = stats.get("mean", float("nan"))
                    sd_v   = stats.get("sd",   0.0)
                    min_v  = stats.get("min",  float("nan"))
                    max_v  = stats.get("max",  float("nan"))
                    ws.cell(row=row, column=1, value=mv_label)
                    ws.cell(row=row, column=2, value=side)
                    ws.cell(row=row, column=3, value=metric_label)
                    ws.cell(row=row, column=4, value=len(values))
                    ws.cell(row=row, column=5,
                            value=round(mean_v, 2) if not _math.isnan(mean_v) else None)
                    ws.cell(row=row, column=6, value=round(sd_v, 2))
                    ws.cell(row=row, column=7,
                            value=round(min_v, 2) if not _math.isnan(min_v) else None)
                    ws.cell(row=row, column=8,
                            value=round(max_v, 2) if not _math.isnan(max_v) else None)
                    row += 1

            chart_fig = self._charts.get(mv_name)
            if chart_fig is not None:
                try:
                    buf = io.BytesIO()
                    chart_fig.savefig(buf, format="png", dpi=120, bbox_inches="tight")
                    buf.seek(0)
                    img = XLImage(buf)
                    img.width  = 480
                    img.height = 260
                    ws.add_image(img, f"J{mv_start}")
                except Exception:
                    pass
            row += 1

    def _xl_write_rep_detail(self, ws, groups: dict) -> None:
        import math as _math
        import numpy as _np
        from openpyxl.styles import Font

        bold = Font(bold=True)
        row = 1

        for mv_name, sides_data in groups.items():
            is_bilateral = len({d.get("c3d_path", "") for _, d in sides_data}) == 1
            mv_label = mv_name + (" (Bilateral)" if is_bilateral else " (Unilateral)")
            ws.cell(row=row, column=1, value=mv_label).font = bold
            row += 1

            all_reps: dict = {}
            for side, data in sides_data:
                reps: list = []
                segments  = data.get("segments", [])
                angle_arr = data.get("angle_data")
                extended  = data.get("extended", {})

                if segments and isinstance(angle_arr, _np.ndarray):
                    for i, (s, e) in enumerate(segments):
                        chunk = angle_arr[s : e + 1]
                        valid = chunk[~_np.isnan(chunk)]
                        if valid.size == 0:
                            reps.append({
                                "rep": i + 1, "peak": None, "peak_frame": None,
                                "valley": None, "valley_frame": None, "rom": None,
                            })
                        else:
                            pk_fr  = s + int(_np.nanargmax(chunk))
                            vl_fr  = s + int(_np.nanargmin(chunk))
                            pk_val = float(_np.nanmax(chunk))
                            vl_val = float(_np.nanmin(chunk))
                            reps.append({
                                "rep":          i + 1,
                                "peak":         round(pk_val, 2),
                                "peak_frame":   pk_fr,
                                "valley":       round(vl_val, 2),
                                "valley_frame": vl_fr,
                                "rom":          round(pk_val - vl_val, 2),
                            })
                else:
                    peak_vals   = extended.get("peak",   {}).get("values", [])
                    valley_vals = extended.get("valley", {}).get("values", [])
                    rom_vals    = extended.get("rom",    {}).get("values", [])
                    n = max(len(peak_vals), len(valley_vals), len(rom_vals), 0)
                    for i in range(n):
                        pk = peak_vals[i]   if i < len(peak_vals)   else float("nan")
                        vl = valley_vals[i] if i < len(valley_vals) else float("nan")
                        rm = rom_vals[i]    if i < len(rom_vals)    else float("nan")
                        reps.append({
                            "rep":          i + 1,
                            "peak":         round(pk, 2) if not _math.isnan(pk) else None,
                            "peak_frame":   None,
                            "valley":       round(vl, 2) if not _math.isnan(vl) else None,
                            "valley_frame": None,
                            "rom":          round(rm, 2) if not _math.isnan(rm) else None,
                        })
                all_reps[side] = reps

            # Column order per side: Peak Frame | Peak Time (s) | Peak (°) |
            #                        Valley Frame | Valley Time (s) | Valley (°) | ROM (°)
            col = 1
            ws.cell(row=row, column=col, value="Rep").font = bold
            col += 1
            for side, _ in sides_data:
                pfx = "right_" if side == "Right" else "left_"
                ws.cell(row=row, column=col,     value=t(pfx + "col_peak_frame")).font   = bold
                ws.cell(row=row, column=col + 1, value=t(pfx + "col_peak_time")).font    = bold
                ws.cell(row=row, column=col + 2, value=t(pfx + "col_peak_angle")).font   = bold
                ws.cell(row=row, column=col + 3, value=t(pfx + "col_valley_frame")).font = bold
                ws.cell(row=row, column=col + 4, value=t(pfx + "col_valley_time")).font  = bold
                ws.cell(row=row, column=col + 5, value=t(pfx + "col_valley_angle")).font = bold
                ws.cell(row=row, column=col + 6, value="ROM (°)").font                   = bold
                col += 7
            row += 1

            max_reps = max((len(r) for r in all_reps.values()), default=0)
            for rep_i in range(max_reps):
                ws.cell(row=row, column=1, value=rep_i + 1)
                col = 2
                for side, _ in sides_data:
                    reps = all_reps.get(side, [])
                    if rep_i < len(reps):
                        r = reps[rep_i]
                        pk_fr = r["peak_frame"]
                        vl_fr = r["valley_frame"]
                        ws.cell(row=row, column=col,     value=pk_fr)
                        ws.cell(row=row, column=col + 1,
                                value=round(pk_fr / 100.0, 2) if pk_fr is not None else None)
                        ws.cell(row=row, column=col + 2, value=r["peak"])
                        ws.cell(row=row, column=col + 3, value=vl_fr)
                        ws.cell(row=row, column=col + 4,
                                value=round(vl_fr / 100.0, 2) if vl_fr is not None else None)
                        ws.cell(row=row, column=col + 5, value=r["valley"])
                        ws.cell(row=row, column=col + 6, value=r["rom"])
                    col += 7
                row += 1
            row += 1

    def _xl_write_raw_data(
        self,
        ws,
        groups: dict,
    ) -> None:
        import math as _math
        import numpy as _np
        from openpyxl.styles import Font
        from openpyxl.chart import LineChart, Reference, Series
        from openpyxl.utils import get_column_letter

        bold = Font(bold=True)
        TITLE_ROW   = 1
        HEADER_ROW  = 2
        DATA_START  = 3
        current_col = 1

        for mv_name, sides_data in groups.items():
            side_arrays: dict = {}
            max_len    = 0
            frame_rate = 100.0
            for side, data in sides_data:
                arr = data.get("angle_data")
                fr  = data.get("frame_rate")
                if fr:
                    frame_rate = float(fr)
                if isinstance(arr, _np.ndarray):
                    side_arrays[side] = arr
                    max_len = max(max_len, len(arr))

            n_sides      = len(sides_data)
            is_bilateral = len({d.get("c3d_path", "") for _, d in sides_data}) == 1
            label_suffix = " (Bilateral)" if is_bilateral else " (Unilateral)"

            block_width = 2 + n_sides

            # Title row
            ws.cell(row=TITLE_ROW, column=current_col,
                    value=mv_name + label_suffix).font = bold

            # Header row
            ws.cell(row=HEADER_ROW, column=current_col,     value="Frame").font = bold
            ws.cell(row=HEADER_ROW, column=current_col + 1, value=t("col_time_s")).font = bold
            for c_i, (side, _) in enumerate(sides_data):
                ws.cell(row=HEADER_ROW, column=current_col + 2 + c_i,
                        value=f"{side} (°)").font = bold

            # Data rows
            sides_order  = [side for side, _ in sides_data]
            data_end_row = DATA_START - 1
            for fr in range(max_len):
                r = DATA_START + fr
                ws.cell(row=r, column=current_col,     value=fr)
                ws.cell(row=r, column=current_col + 1, value=round(fr / frame_rate, 4))
                for c_i, side in enumerate(sides_order):
                    arr = side_arrays.get(side)
                    if arr is not None and fr < len(arr):
                        val = float(arr[fr])
                        ws.cell(row=r, column=current_col + 2 + c_i,
                                value=None if _math.isnan(val) else round(val, 4))
                data_end_row = r

            # LineChart
            if max_len > 0:
                chart = LineChart()
                chart.title        = mv_name
                chart.x_axis.title = t("col_time_s")
                chart.width  = 15
                chart.height = 10

                x_ref = Reference(ws, min_col=current_col + 1,
                                  max_col=current_col + 1,
                                  min_row=DATA_START, max_row=data_end_row)
                for c_i, side in enumerate(sides_order):
                    y_ref = Reference(ws, min_col=current_col + 2 + c_i,
                                      max_col=current_col + 2 + c_i,
                                      min_row=DATA_START, max_row=data_end_row)
                    series_title = (t("left_side") if side != "Right"
                                    else t("right_side"))
                    s = Series(y_ref, title=series_title)
                    chart.append(s)
                chart.set_categories(x_ref)
                anchor_col = get_column_letter(current_col)
                ws.add_chart(chart, f"{anchor_col}{data_end_row + 2}")

            current_col += block_width + 1

    def _export_chart(self, fig, name: str = "chart") -> None:
        path = filedialog.asksaveasfilename(
            title=t("export_chart_dialog_title"),
            defaultextension=".png",
            filetypes=[("PNG", "*.png"), ("All files", "*.*")],
            initialfile=f"{name or 'chart'}.png",
        )
        if path:
            fig.savefig(path, dpi=150, bbox_inches="tight")

    def _generate_report(self) -> None:
        if not self._processed:
            return

        out_path = filedialog.asksaveasfilename(
            title=t("s4_excel_dialog_title"),
            defaultextension=".xlsx",
            filetypes=[("Excel Workbook", "*.xlsx"), ("All files", "*.*")],
            initialfile="ROM_Report.xlsx",
        )
        if not out_path:
            return

        try:
            from excel_export import export_excel
            export_excel(self._processed, out_path)
        except Exception as exc:
            messagebox.showerror("Export error", str(exc))
            return

        messagebox.showinfo(
            t("s4_excel_saved_title"),
            t("s4_excel_saved_msg").format(path=out_path),
        )

    def _new_analysis(self) -> None:
        self._import_rows.clear()
        self._processed.clear()
        self._process_btn = None
        self._laterality_var.set("unilateral")
        self._recording_type_var.set("continuous")
        self._num_reps_var.set(6)
        self._layout_vertical = True
        self._show_screen_1()


# ══════════════════════════════════════════════════════════════════════════
#  _TrunkSegmentationWindow — segmentation modal for computed trunk angles
# ══════════════════════════════════════════════════════════════════════════

class _TrunkSegmentationWindow(ctk.CTkToplevel):
    """
    Segmentation modal for Trunk Lateral Inclination.

    Shows all three trunk angle components (via plot_trunk_inclination) as the
    main visualisation while running segmentation on the lateral_inclination
    1D array.  Supports Auto, Manual, and Events modes — identical behaviour
    to C3DSegmentationWindow.

    After accept, ``self.result`` holds the same dict structure as
    C3DSegmentationWindow (segments, roms, mean, sd, outlier_flags).
    """

    def __init__(
        self,
        parent,
        movement_name: str,
        angle_data: np.ndarray,
        trunk_angles: dict,
        frame_rate: int,
        events: list[dict] | None = None,
    ) -> None:
        super().__init__(parent)
        self.title(f"{t('seg_window_title')} — {movement_name}")
        self.geometry("1060x820")
        self.grab_set()
        self.lift()
        self.focus_force()

        import numpy as _np
        from matplotlib.backends.backend_tkagg import (
            FigureCanvasTkAgg, NavigationToolbar2Tk,
        )
        from plotting import plot_trunk_inclination

        self._angle_data  = angle_data.astype(float)
        self._trunk_angles = trunk_angles
        self._frame_rate  = frame_rate
        self._events      = events or []
        self._movement_name = movement_name

        self.result: dict | None = None
        self.go_back: bool = False
        self.last_direction: str = "peak_to_valley"

        # ── Segmentation state ────────────────────────────────────────────
        self._segments: list[tuple[int, int]] = []
        self._peaks: _np.ndarray = _np.array([], dtype=int)
        self._valleys: _np.ndarray = _np.array([], dtype=int)
        self._markers: list[float] = []
        self._excluded_indices: set[int] = set()
        self._mode_var = ctk.StringVar(value="auto")

        # ── Trunk inclination figure (single panel — lateral inclination) ──
        self._trunk_fig = plot_trunk_inclination(
            trunk_angles, frame_rate, title=movement_name
        )

        # Single-panel figure: axes[0] is the lateral inclination subplot
        # where segmentation overlays (bands, peak/valley markers) are drawn.
        self._lat_ax = self._trunk_fig.axes[0]

        self._canvas = FigureCanvasTkAgg(self._trunk_fig, master=self)
        self._canvas.get_tk_widget().pack(
            fill="both", expand=True, padx=10, pady=(10, 0))

        nav_frame = ctk.CTkFrame(self, height=30)
        nav_frame.pack(fill="x", padx=10)
        nav_frame.pack_propagate(False)
        nav = NavigationToolbar2Tk(self._canvas, nav_frame)
        nav.update()

        self._cid = self._canvas.mpl_connect(
            "button_press_event", self._on_canvas_click)

        # ── Mode selector ─────────────────────────────────────────────────
        mode_bar = ctk.CTkFrame(self)
        mode_bar.pack(fill="x", padx=10, pady=(6, 0))
        ctk.CTkLabel(
            mode_bar, text=t("seg_mode_label"),
            font=ctk.CTkFont(weight="bold"),
        ).pack(side="left", padx=(8, 4))
        for txt, val in [
            (t("seg_mode_auto"),   "auto"),
            (t("seg_mode_manual"), "manual"),
            (t("seg_mode_events"), "events"),
        ]:
            ctk.CTkRadioButton(
                mode_bar, text=txt, variable=self._mode_var, value=val,
                command=self._switch_mode,
            ).pack(side="left", padx=6)

        # ── Mode panels ───────────────────────────────────────────────────
        self._panel_container = ctk.CTkFrame(self, height=155)
        self._panel_container.pack(fill="x", padx=10, pady=4)
        self._panel_container.pack_propagate(False)

        self._panel_auto   = self._build_auto_panel(self._panel_container)
        self._panel_manual = self._build_manual_panel(self._panel_container)
        self._panel_events = self._build_events_panel(self._panel_container)

        # ── Stats label ───────────────────────────────────────────────────
        stats_frame = ctk.CTkFrame(self)
        stats_frame.pack(fill="x", padx=10, pady=(0, 4))
        self._stats_lbl = ctk.CTkLabel(
            stats_frame,
            text=f"  {t('seg_no_segments')}",
            font=ctk.CTkFont(size=11), justify="left",
        )
        self._stats_lbl.pack(anchor="w", padx=8, pady=4)

        # ── Action buttons ────────────────────────────────────────────────
        btn_bar = ctk.CTkFrame(self)
        btn_bar.pack(fill="x", padx=10, pady=(0, 10))
        ctk.CTkButton(
            btn_bar, text=t("seg_accept"),
            fg_color="#2D7A2D", hover_color="#1F5C1F",
            font=ctk.CTkFont(weight="bold"), width=160,
            command=self._accept,
        ).pack(side="right", padx=4)
        ctk.CTkButton(
            btn_bar, text=t("seg_cancel"), width=120,
            command=self.destroy,
        ).pack(side="right")

        self._switch_mode()

    # ── Panel builders ─────────────────────────────────────────────────────

    def _build_auto_panel(self, parent) -> ctk.CTkFrame:
        from config import TRUNK_LATERAL_MIN_PROMINENCE, TRUNK_LATERAL_MIN_DISTANCE

        pnl = ctk.CTkFrame(parent, fg_color="transparent")

        self._prom_var  = ctk.IntVar(value=TRUNK_LATERAL_MIN_PROMINENCE)
        self._dist_var  = ctk.IntVar(value=TRUNK_LATERAL_MIN_DISTANCE)
        self._cycle_var = ctk.StringVar(value="halfcycle")

        row1 = ctk.CTkFrame(pnl, fg_color="transparent")
        row1.pack(anchor="w", padx=8, pady=4)
        ctk.CTkLabel(row1, text=t("seg_prominence"), width=130, anchor="w").pack(side="left")
        ctk.CTkSlider(row1, from_=1, to=60, variable=self._prom_var, width=180).pack(side="left", padx=4)
        ctk.CTkLabel(row1, textvariable=self._prom_var, width=30).pack(side="left")

        row2 = ctk.CTkFrame(pnl, fg_color="transparent")
        row2.pack(anchor="w", padx=8, pady=4)
        ctk.CTkLabel(row2, text=t("seg_min_distance"), width=130, anchor="w").pack(side="left")
        ctk.CTkSlider(row2, from_=10, to=300, variable=self._dist_var, width=180).pack(side="left", padx=4)
        ctk.CTkLabel(row2, textvariable=self._dist_var, width=36).pack(side="left")

        row3 = ctk.CTkFrame(pnl, fg_color="transparent")
        row3.pack(anchor="w", padx=8, pady=2)
        ctk.CTkLabel(row3, text=t("seg_cycle_from"), width=130, anchor="w").pack(side="left")
        for lbl, val in [
            (t("seg_valley_valley"), "valley"),
            (t("seg_peak_peak"),     "peak"),
        ]:
            ctk.CTkRadioButton(
                row3, text=lbl, variable=self._cycle_var, value=val,
            ).pack(side="left", padx=4)

        ctk.CTkButton(pnl, text=t("seg_detect"), width=100,
                      command=self._auto_detect).pack(anchor="w", padx=8, pady=4)
        return pnl

    def _build_manual_panel(self, parent) -> ctk.CTkFrame:
        pnl = ctk.CTkFrame(parent, fg_color="transparent")
        ctk.CTkLabel(
            pnl, text=t("seg_manual_hint"),
            font=ctk.CTkFont(size=10), text_color="gray",
        ).pack(anchor="w", padx=8, pady=4)
        row = ctk.CTkFrame(pnl, fg_color="transparent")
        row.pack(anchor="w", padx=8)
        ctk.CTkButton(row, text=t("seg_undo"), width=80,
                      command=self._manual_undo).pack(side="left", padx=4)
        ctk.CTkButton(row, text=t("seg_clear"), width=80,
                      command=self._manual_clear).pack(side="left", padx=4)
        self._manual_lbl = ctk.CTkLabel(
            pnl, text=f"  {t('seg_initial_markers')}",
            font=ctk.CTkFont(size=10),
        )
        self._manual_lbl.pack(anchor="w", padx=8, pady=2)
        return pnl

    def _build_events_panel(self, parent) -> ctk.CTkFrame:
        pnl = ctk.CTkFrame(parent, fg_color="transparent")
        event_names = [e["name"] for e in self._events]

        if not event_names:
            ctk.CTkLabel(pnl, text=t("seg_no_events"),
                         font=ctk.CTkFont(size=10), text_color="gray").pack(
                anchor="w", padx=8, pady=8)
            return pnl

        self._ev_start_var = ctk.StringVar(value=event_names[0])
        self._ev_end_var   = ctk.StringVar(
            value=event_names[1] if len(event_names) > 1 else event_names[0])

        row = ctk.CTkFrame(pnl, fg_color="transparent")
        row.pack(anchor="w", padx=8, pady=6)
        ctk.CTkLabel(row, text=t("seg_start_event"), width=100, anchor="w").pack(side="left")
        ctk.CTkOptionMenu(row, values=event_names,
                          variable=self._ev_start_var, width=150).pack(side="left", padx=4)
        ctk.CTkLabel(row, text=t("seg_end_event"), width=90, anchor="w").pack(side="left")
        ctk.CTkOptionMenu(row, values=event_names,
                          variable=self._ev_end_var, width=150).pack(side="left", padx=4)
        ctk.CTkButton(pnl, text=t("seg_map_events"), width=110,
                      command=self._map_events).pack(anchor="w", padx=8, pady=4)
        return pnl

    # ── Mode switching ─────────────────────────────────────────────────────

    def _switch_mode(self) -> None:
        mode = self._mode_var.get()
        for pnl in (self._panel_auto, self._panel_manual, self._panel_events):
            pnl.pack_forget()
        if mode == "auto":
            self._panel_auto.pack(fill="x")
        elif mode == "manual":
            self._panel_manual.pack(fill="x")
        else:
            self._panel_events.pack(fill="x")
        self._redraw()

    # ── Auto detection ─────────────────────────────────────────────────────

    def _auto_detect(self) -> None:
        from segmentation import auto_segment, compute_rep_roms
        try:
            segs, peaks, valleys = auto_segment(
                self._angle_data,
                min_prominence=float(self._prom_var.get()),
                min_distance=int(self._dist_var.get()),
                cycle_from=self._cycle_var.get(),
            )
        except Exception as exc:
            messagebox.showwarning(t("seg_detect_error"), str(exc), parent=self)
            return
        self._segments = segs
        self._peaks    = peaks
        self._valleys  = valleys
        self._excluded_indices.clear()
        self._redraw()
        self._update_stats_label()

    # ── Manual mode ───────────────────────────────────────────────────────

    def _on_canvas_click(self, event) -> None:
        if self._mode_var.get() != "manual":
            return
        if event.inaxes is None:
            return

        # Identify if click is on the lateral_inclination subplot
        if event.inaxes not in self._trunk_fig.axes:
            return

        x_frame = event.xdata * self._frame_rate  # convert time → frames

        if event.button == 1:
            self._markers.append(x_frame)
        elif event.button == 3 and self._markers:
            nearest = min(range(len(self._markers)),
                          key=lambda i: abs(self._markers[i] - x_frame))
            self._markers.pop(nearest)

        pairs = len(self._markers) // 2
        self._segments = [
            (int(self._markers[2 * i]), int(self._markers[2 * i + 1]))
            for i in range(pairs)
        ]
        self._manual_lbl.configure(
            text=t("seg_markers_placed").format(
                n=len(self._markers), pairs=pairs))
        self._redraw()
        self._update_stats_label()

    def _manual_undo(self) -> None:
        if self._markers:
            self._markers.pop()
            pairs = len(self._markers) // 2
            self._segments = [
                (int(self._markers[2 * i]), int(self._markers[2 * i + 1]))
                for i in range(pairs)
            ]
            self._redraw()
            self._update_stats_label()

    def _manual_clear(self) -> None:
        self._markers.clear()
        self._segments.clear()
        self._manual_lbl.configure(text=f"  {t('seg_initial_markers')}")
        self._redraw()
        self._update_stats_label()

    # ── Events mode ───────────────────────────────────────────────────────

    def _map_events(self) -> None:
        if not self._events:
            return
        start_name = self._ev_start_var.get()
        end_name   = self._ev_end_var.get()
        starts = [e["frame"] for e in self._events if e["name"] == start_name]
        ends   = [e["frame"] for e in self._events if e["name"] == end_name]
        pairs  = [(s, e) for s in starts for e in ends if e > s]
        if not pairs:
            messagebox.showwarning(
                t("seg_no_pairs_title"),
                t("seg_no_pairs_msg").format(start=start_name, end=end_name),
                parent=self,
            )
            return
        # Keep only non-overlapping pairs
        segs: list[tuple[int, int]] = []
        last_end = -1
        for s, e in sorted(pairs):
            if s > last_end:
                segs.append((s, e))
                last_end = e
        self._segments = segs
        self._excluded_indices.clear()
        self._redraw()
        self._update_stats_label()

    # ── Drawing ────────────────────────────────────────────────────────────

    def _redraw(self) -> None:
        import numpy as _np
        from segmentation import compute_rep_roms

        # Clear only segmentation overlays from the lateral inclination subplot
        lat_ax = self._lat_ax
        for artist in list(lat_ax.collections) + list(lat_ax.lines[1:]):
            try:
                artist.remove()
            except Exception:
                pass

        scale = 1.0 / self._frame_rate
        lat_data = self._angle_data

        # Draw segments as shaded bands
        _REP_COLORS = [
            "#FF6B6B", "#4ECDC4", "#45B7D1", "#96CEB4",
            "#FFEAA7", "#DDA0DD", "#98D8C8",
        ]
        roms = compute_rep_roms(lat_data, self._segments) if self._segments else []
        for i, (s, e) in enumerate(self._segments):
            color = _REP_COLORS[i % len(_REP_COLORS)]
            xs, xe = s * scale, e * scale
            alpha = 0.10 if i in self._excluded_indices else 0.22
            lat_ax.axvspan(xs, xe, alpha=alpha, color=color, zorder=1)
            label = f"R{i + 1}"
            if roms and i < len(roms) and not _np.isnan(roms[i]):
                label += f"\n{roms[i]:.1f}°"
            lat_ax.text(
                (xs + xe) / 2, 0.97, label,
                fontsize=7, ha="center", va="top",
                transform=lat_ax.get_xaxis_transform(),
                color=color,
            )

        # Manual markers as vertical lines
        if self._mode_var.get() == "manual":
            for j, m in enumerate(self._markers):
                color = "#2D7A2D" if j % 2 == 0 else "#E05252"
                lat_ax.axvline(m * scale, color=color, linewidth=1.2,
                               linestyle="--", zorder=3)

        # Auto peaks/valleys
        if self._mode_var.get() == "auto" and len(self._peaks):
            n = len(lat_data)
            lat_ax.plot(
                self._peaks * scale,
                lat_data[_np.clip(self._peaks, 0, n - 1)],
                "v", color="#E05252", ms=5, zorder=5,
            )
            lat_ax.plot(
                self._valleys * scale,
                lat_data[_np.clip(self._valleys, 0, n - 1)],
                "^", color="#2D7A2D", ms=5, zorder=5,
            )

        self._canvas.draw_idle()

    def _update_stats_label(self) -> None:
        from segmentation import compute_rep_roms, compute_stats_from_roms
        active = [s for i, s in enumerate(self._segments)
                  if i not in self._excluded_indices]
        if not active:
            self._stats_lbl.configure(text=f"  {t('seg_no_segments')}")
            return
        roms  = compute_rep_roms(self._angle_data, active)
        stats = compute_stats_from_roms(roms, active)
        txt = (
            f"  {t('seg_n_segments').format(n=len(active))}   "
            f"Mean ROM: {stats['mean']:.1f}°   SD: {stats['sd']:.1f}°"
        )
        self._stats_lbl.configure(text=txt)

    # ── Accept / Cancel ────────────────────────────────────────────────────

    def _accept(self) -> None:
        active_segs = [s for i, s in enumerate(self._segments)
                       if i not in self._excluded_indices]
        if not active_segs:
            messagebox.showwarning(
                t("seg_accept_warn_title"),
                t("seg_accept_warn_msg"),
                parent=self,
            )
            return
        from segmentation import compute_rep_roms, compute_stats_from_roms
        roms  = compute_rep_roms(self._angle_data, active_segs)
        stats = compute_stats_from_roms(roms, active_segs)
        self.result = stats
        self.destroy()
